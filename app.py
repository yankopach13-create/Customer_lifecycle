"""
Цикл жизни клиента в продукте.
Streamlit-приложение для загрузки отчётов из Qlik по шаблону.
"""

import base64
import io
import json
import re
import streamlit as st
from openpyxl.styles import Alignment, Font
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots

# Структура документа: столбец 0 — категория/продукт, 1—2 — период, 3 — количество, 4 — код клиента
COL_CATEGORY = "category"
COL_PERIOD_MAIN = "period_main"
COL_PERIOD_SUB = "period_sub"
COL_QUANTITY = "quantity"
COL_CLIENT = "client_id"

# 8 поведенческих кластеров: порядок по убыванию активности, название → описание
CLUSTER_8_ORDER = [
    "Активные (VIP)",
    "Регулярные с высоким объёмом",
    "Крупные нерегулярные",
    "Средняя активность",
    "Периодические (малый объём)",
    "Низкая активность",
    "Разовая крупная покупка",
    "Разовая покупка",
]
CLUSTER_8_DESCRIPTIONS = {
    "Активные (VIP)": "Высокий объём покупок и высокая регулярность: покупают часто и много.",
    "Регулярные с высоким объёмом": "Высокий объём и стабильная регулярность: постоянные крупные покупатели.",
    "Крупные нерегулярные": "Высокий объём, но покупают не в каждый период: крупные, но редкие покупки.",
    "Средняя активность": "Средний объём и средняя регулярность: умеренная вовлечённость.",
    "Периодические (малый объём)": "Низкий объём, но покупают часто: стабильные малые покупки.",
    "Низкая активность": "Низкий объём и невысокая регулярность: эпизодические малые покупки.",
    "Разовая крупная покупка": "Высокий объём за один или два периода: разовая крупная сделка.",
    "Разовая покупка": "Низкий объём и одна-две покупки: попробовали продукт.",
    "Не покупали": "В выбранном окне не покупали анализируемый продукт.",
}


def _cluster_display_name(name: str) -> str:
    """Убирает всё в скобках из названия кластера для отображения."""
    return re.sub(r"\s*\([^)]*\)", "", name).strip() if name else name


def create_copy_button(text: str, button_label: str, key: str) -> None:
    """Создаёт кнопку для копирования текста в буфер обмена (Clipboard API + fallback)."""
    safe_key = re.sub(r"[^a-zA-Z0-9_]", "_", str(key))
    text_json = json.dumps(text)
    html = f"""
    <div data-testid="stButton" style="width: 100%; margin: 5px 0;">
        <button id="copy_btn_{safe_key}" onclick="copyToClipboard_{safe_key}()" style="
            width: 100%;
            padding: 12px 16px;
            background: transparent !important;
            color: #fff !important;
            font-weight: 700 !important;
            border: 2px solid #adb5bd !important;
            border-radius: 8px !important;
            cursor: pointer !important;
            font-weight: 400 !important;
            font-size: 0.85rem !important;
            line-height: 1.3 !important;
            text-align: center !important;
            min-height: 72px !important;
            height: 72px !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            white-space: normal !important;
            word-wrap: break-word !important;
            overflow-wrap: break-word !important;
            box-shadow: none !important;
            transition: all 0.3s ease !important;
            margin: 0 !important;
            box-sizing: border-box !important;
            position: relative !important;
        " onmouseover="if (!this.classList.contains('copied')) {{ this.style.transform='translateY(-2px)'; this.style.boxShadow='0 2px 8px rgba(0,0,0,0.08)'; this.style.borderColor='#6c757d'; }}" onmouseout="if (!this.classList.contains('copied')) {{ this.style.transform='translateY(0)'; this.style.boxShadow='none'; this.style.borderColor='#adb5bd'; }}" onmousedown="if (!this.classList.contains('copied')) {{ this.style.transform='translateY(0)'; }}" onmouseup="if (!this.classList.contains('copied')) {{ this.style.transform='translateY(-2px)'; }}">
            <div style="display: flex; align-items: center; justify-content: center; width: 100%;">
                <p id="copy_btn_text_{safe_key}" style="margin: 0; padding: 0; font-size: 0.85rem; font-weight: 700; color: #fff; line-height: 1.3; word-wrap: break-word; overflow-wrap: break-word; white-space: normal;">{button_label}</p>
            </div>
        </button>
    </div>
    <script>
        const textToCopy_{safe_key} = {text_json};
        function copyToClipboard_{safe_key}() {{
            const text = textToCopy_{safe_key};
            const button = document.getElementById('copy_btn_{safe_key}');
            const buttonText = document.getElementById('copy_btn_text_{safe_key}');
            const originalText = buttonText.innerHTML;
            function showSuccess() {{
                button.classList.add('copied');
                button.style.background = 'linear-gradient(135deg, #4CAF50 0%, #45a049 100%)';
                button.style.borderColor = '#4CAF50';
                button.style.color = 'white';
                button.style.transform = 'scale(0.98)';
                buttonText.innerHTML = '✓ Скопировано!';
                setTimeout(function() {{
                    button.classList.remove('copied');
                    button.style.background = 'transparent';
                    button.style.borderColor = '#adb5bd';
                    button.style.color = '#fff';
                    button.style.transform = 'translateY(0)';
                    buttonText.innerHTML = originalText;
                }}, 2500);
            }}
            if (navigator.clipboard && navigator.clipboard.writeText) {{
                navigator.clipboard.writeText(text).then(function() {{ showSuccess(); }}).catch(function(err) {{
                    console.error('Clipboard API error:', err);
                    fallbackCopy_{safe_key}(text, showSuccess);
                }});
            }} else {{
                fallbackCopy_{safe_key}(text, showSuccess);
            }}
        }}
        function fallbackCopy_{safe_key}(text, successCallback) {{
            const textarea = document.createElement('textarea');
            textarea.value = text;
            textarea.style.position = 'fixed';
            textarea.style.left = '-999999px';
            textarea.style.top = '-999999px';
            textarea.style.opacity = '0';
            document.body.appendChild(textarea);
            textarea.focus();
            textarea.select();
            try {{
                const successful = document.execCommand('copy');
                if (successful) {{ successCallback(); }}
                else {{ alert('Не удалось скопировать. Скопируйте вручную.'); }}
            }} catch(err) {{
                console.error('Copy command error:', err);
                alert('Ошибка копирования: ' + err);
            }} finally {{
                document.body.removeChild(textarea);
            }}
        }}
    </script>
    """
    components.html(html, height=85)


def create_excel_download_button(excel_bytes: bytes, filename: str, button_label: str, key: str) -> None:
    """Создаёт HTML-кнопку скачивания Excel (полный контроль над размером и визуалом, как на скриншоте)."""
    safe_key = re.sub(r"[^a-zA-Z0-9_]", "_", str(key))
    b64 = base64.b64encode(excel_bytes).decode("ascii")
    filename_esc = json.dumps(filename)
    html = f"""
    <div style="width: 100%; margin: 0;">
        <button id="excel_btn_{safe_key}" type="button" style="
            width: 100%;
            height: 2.375rem;
            min-height: 2.375rem;
            padding: 6px 12px;
            background: transparent;
            color: white;
            font-weight: 700;
            font-size: 0.85rem;
            border: 2px solid white;
            border-radius: 6px;
            cursor: pointer;
            text-align: center;
            line-height: 1.2;
            box-sizing: border-box;
            box-shadow: 0 2px 8px rgba(0,0,0,0.2);
            transition: transform 0.15s ease, box-shadow 0.2s ease, background 0.2s ease;
            user-select: none;
        " onmouseover="this.style.transform='scale(1.02)'; this.style.boxShadow='0 4px 14px rgba(0,0,0,0.25), 0 0 0 1px rgba(255,255,255,0.15)'; this.style.background='rgba(255,255,255,0.08)';" onmouseout="this.style.transform='scale(1)'; this.style.boxShadow='0 2px 8px rgba(0,0,0,0.2)'; this.style.background='transparent';" onmousedown="this.style.transform='scale(0.98)';" onmouseup="this.style.transform='scale(1.02)';" onmouseleave="this.style.transform='scale(1)';">
            {button_label}
        </button>
    </div>
    <script>
        (function() {{
            var btn = document.getElementById('excel_btn_{safe_key}');
            var b64 = {json.dumps(b64)};
            var filename = {filename_esc};
            btn.addEventListener('click', function() {{
                var dataUri = 'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,' + b64;
                var a = document.createElement('a');
                a.href = dataUri;
                a.download = filename;
                document.body.appendChild(a);
                a.click();
                document.body.removeChild(a);
            }});
        }})();
    </script>
    """
    components.html(html, height=46)


def _norm_client_id(ser: pd.Series) -> pd.Series:
    """Приводит коды клиентов к одному строковому виду (348385 и 348385.0 → одинаково)."""
    s = ser.astype(str).str.strip()
    # убираем хвост .0 у целых чисел, чтобы совпадали коды из разных файлов
    return s.str.replace(r"\.0$", "", regex=True)


def load_and_normalize(uploaded_file):
    """Читает Excel и приводит к стандартным столбцам: category, period_main, period_sub, quantity, client_id."""
    if uploaded_file is None:
        return None
    raw = pd.read_excel(uploaded_file, header=0)
    # Берём первые 5 столбцов по позиции
    cols = raw.iloc[:, :5].copy()
    cols.columns = [COL_CATEGORY, COL_PERIOD_MAIN, COL_PERIOD_SUB, COL_QUANTITY, COL_CLIENT]
    # Убираем строки с пустыми ключевыми полями
    cols = cols.dropna(subset=[COL_PERIOD_MAIN, COL_CLIENT])
    cols[COL_QUANTITY] = pd.to_numeric(cols[COL_QUANTITY], errors="coerce").fillna(0).astype(int)
    cols[COL_CATEGORY] = cols[COL_CATEGORY].astype(str).str.strip()
    cols[COL_CLIENT] = _norm_client_id(cols[COL_CLIENT])
    return cols


def merge_and_prepare(df1, df2):
    """Объединяет два документа и готовит период, порядок периодов и когорту (первый период клиента)."""
    df = pd.concat([df1, df2], ignore_index=True)
    df[COL_PERIOD_MAIN] = df[COL_PERIOD_MAIN].astype(str).str.strip()
    df[COL_PERIOD_SUB] = df[COL_PERIOD_SUB].astype(str).str.strip()
    # Порядок периодов для определения когорты
    period_order = (
        df[[COL_PERIOD_MAIN, COL_PERIOD_SUB]]
        .drop_duplicates()
        .sort_values([COL_PERIOD_MAIN, COL_PERIOD_SUB])
        .reset_index(drop=True)
    )
    period_order["period_rank"] = period_order.index
    df = df.merge(
        period_order,
        on=[COL_PERIOD_MAIN, COL_PERIOD_SUB],
        how="left",
    )
    first_rank = df.groupby(COL_CLIENT)["period_rank"].min().rename("first_period_rank")
    df = df.merge(first_rank, left_on=COL_CLIENT, right_index=True, how="left")
    rank_to_period = period_order.set_index("period_rank")[[COL_PERIOD_MAIN, COL_PERIOD_SUB]]
    return df, period_order, rank_to_period, first_rank


def format_period_short(period_main, period_sub):
    """Форматирует период как 25/1, 25/2 (год/неделя)."""
    pm, ps = str(period_main).strip(), str(period_sub).strip()
    year_match = re.search(r"20\d{2}|\d{4}", pm)
    year_short = year_match.group(0)[-2:] if year_match else (pm[-2:] if len(pm) >= 2 else "")
    week_match = re.search(r"\d+", ps)
    week = week_match.group(0) if week_match else ps
    return f"{year_short}/{week}" if year_short and week else f"{pm} {ps}".strip()


# Сокращения месяцев для строки «По данным за янв-фев 2025»
_MONTH_ABBR = {
    "январь": "янв", "февраль": "фев", "март": "мар", "апрель": "апр",
    "май": "май", "июнь": "июн", "июль": "июл", "август": "авг",
    "сентябрь": "сен", "октябрь": "окт", "ноябрь": "ноя", "декабрь": "дек",
}


def format_period_range_for_caption(cohorts_to_use, cohort_ranks, rank_to_period, k_periods, is_months):
    """
    Формирует подпись диапазона периодов: «По данным за 1-6 недель 2025» или «По данным за янв-фев 2025».
    """
    if not cohorts_to_use or not rank_to_period.index.size:
        return ""
    r_min = min(cohort_ranks[lb] for lb in cohorts_to_use)
    r_max = max(cohort_ranks[lb] for lb in cohorts_to_use) + int(k_periods) - 1
    r_max = min(r_max, rank_to_period.index.max())
    r_min = max(r_min, rank_to_period.index.min())
    first = rank_to_period.loc[r_min]
    last = rank_to_period.loc[r_max]
    pm_f, ps_f = str(first[COL_PERIOD_MAIN]).strip(), str(first[COL_PERIOD_SUB]).strip()
    pm_l, ps_l = str(last[COL_PERIOD_MAIN]).strip(), str(last[COL_PERIOD_SUB]).strip()
    year_match = re.search(r"20\d{2}|\d{4}", pm_f)
    year = year_match.group(0) if year_match else (pm_l if re.search(r"\d{4}", pm_l) else "")
    if is_months:
        abbr = lambda s: _MONTH_ABBR.get(s.lower(), s[:3].lower() if len(s) >= 3 else s)
        part = f"{abbr(ps_f)}-{ps_l}" if ps_f != ps_l else abbr(ps_f)
        return f"По данным за {part} {year}"
    w_f = re.search(r"\d+", ps_f)
    w_l = re.search(r"\d+", ps_l)
    week_f = w_f.group(0) if w_f else ps_f
    week_l = w_l.group(0) if w_l else ps_l
    return f"По данным за {week_f}-{week_l} недель {year}"


def _html_to_plain_text(html: str) -> str:
    """Убирает HTML-теги и лишние пробелы, оставляет читаемый текст."""
    if not html:
        return ""
    text = re.sub(r"<[^>]+>", " ", html)
    text = re.sub(r"&nbsp;", " ", text)
    text = re.sub(r"&amp;", "&", text)
    text = re.sub(r"&lt;", "<", text)
    text = re.sub(r"&gt;", ">", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _html_to_plain_fragment(html: str) -> str:
    """Убирает HTML-теги из фрагмента, декодирует сущности, нормализует пробелы (одна строка)."""
    if not html:
        return ""
    text = re.sub(r"<[^>]+>", " ", html)
    text = re.sub(r"&nbsp;", " ", text)
    text = re.sub(r"&amp;", "&", text)
    text = re.sub(r"&lt;", "<", text)
    text = re.sub(r"&gt;", ">", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _strip_css_from_html(html: str) -> str:
    """Удаляет блок <style>...</style>, чтобы в текст не попадали стили."""
    if not html:
        return html
    return re.sub(r"<style[^>]*>.*?</style>", "", html, flags=re.DOTALL | re.IGNORECASE)


def build_excel_report(
    cohort_start: str,
    cohort_end: str,
    anchor_product: str,
    categories: list,
    k_periods: int,
    is_months: bool,
    cluster_summary: pd.DataFrame,
    cluster_comments: dict,
    lifecycle_clusters: list,
    lifecycle_table: pd.DataFrame,
    lifecycle_output_text: str,
    lifecycle_output_rows: list = None,
) -> bytes:
    """
    Собирает полный отчёт в Excel: лист 1 — параметры и кластерный анализ (с примечаниями на кластерах),
    лист 2 — цикл жизни (кластеры, таблица, вывод в объединённой ячейке с переносом).
    Метрики в % выводятся в формате процента.
    """
    buffer = io.BytesIO()
    period_word = "месяцев" if is_months else "недель"
    cluster_comments = cluster_comments or {}

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Лист 1: параметры + кластерный анализ
        params_rows = [
            ["Параметры отчёта", ""],
            ["С когорты", cohort_start],
            ["По когорту", cohort_end],
            ["Якорный продукт когорт", anchor_product or "—"],
            ["Анализируемый продукт", ", ".join(categories) if categories else "—"],
            ["Период (недель/месяцев с когорты)", f"{k_periods} {period_word}"],
        ]
        params_df = pd.DataFrame(params_rows)
        params_df.to_excel(writer, sheet_name="Параметры и кластеры", index=False, header=False)

        start_row = len(params_rows) + 2
        if cluster_summary is not None and not cluster_summary.empty:
            export_cols = [c for c in ["cluster", "clients", "pct", "total_volume", "pct_volume", "avg_client_per_period", "avg_regularity"] if c in cluster_summary.columns]
            cluster_export = cluster_summary[export_cols].copy()
            # Присутствие и Регулярность — в том же формате, что и в таблице в программе
            period_word_plural = "месяцев" if is_months else "недель"
            days_per_period = 30 if is_months else 7
            k_int = int(k_periods)
            window_days = k_int * days_per_period

            def _presence_text(avg_r):
                if pd.isna(avg_r):
                    avg_r = 0.0
                x_per = round(avg_r * k_int, 1)
                y_pct = round(avg_r * 100, 1)
                return f"{x_per} {period_word_plural} из {k_int} ({y_pct}%)"

            def _regularity_text(avg_r):
                if pd.isna(avg_r) or avg_r <= 0.001:
                    return "Приходят редко или одна покупка"
                z_days = max(1, round(days_per_period / avg_r))
                suffix = " (вероятно реже)" if z_days >= window_days else ""
                return f"В среднем каждые {z_days} дн.{suffix}"

            cluster_export["presence"] = cluster_export["avg_regularity"].apply(_presence_text)
            cluster_export["regularity"] = cluster_export["avg_regularity"].apply(_regularity_text)
            cluster_export = cluster_export.drop(columns=["avg_regularity"], errors="ignore")

            # Названия кластеров в Excel — как в программе (с отображаемым именем без скобок, кроме «Итого»)
            cluster_export["cluster"] = cluster_export["cluster"].apply(
                lambda x: "Итого" if x == "Итого" else _cluster_display_name(str(x))
            )
            # Метрики в % — значением в %
            if "pct" in cluster_export.columns:
                cluster_export["pct"] = cluster_export["pct"].apply(lambda x: f"{float(x):.1f}%")
            if "pct_volume" in cluster_export.columns:
                cluster_export["pct_volume"] = cluster_export["pct_volume"].apply(lambda x: f"{float(x):.1f}%")
            col_names_ru = {
                "cluster": "Кластер",
                "clients": "Клиентов",
                "pct": "% клиентов",
                "total_volume": "Объём за период",
                "pct_volume": "% объёма",
                "avg_client_per_period": "Средний объём на клиента за период покупки",
                "presence": "Присутствие",
                "regularity": "Регулярность",
            }
            cluster_export = cluster_export.rename(columns=col_names_ru)
            cluster_export.to_excel(writer, sheet_name="Параметры и кластеры", index=False, startrow=start_row)

            # Под таблицей — описание кластеров в 3 столбцах (кластер, описание, критерий), ячейки с переносом текста
            desc_rows = []
            for cluster_name in cluster_summary["cluster"].tolist():
                if cluster_name == "Итого":
                    continue
                comment_text = cluster_comments.get(cluster_name, "")
                if "\n\nКритерии: " in comment_text:
                    desc_part, crit_part = comment_text.split("\n\nКритерии: ", 1)
                    desc_rows.append((_cluster_display_name(cluster_name), desc_part.strip(), crit_part.strip()))
                else:
                    desc_rows.append((_cluster_display_name(cluster_name), comment_text.strip(), ""))
            if desc_rows:
                desc_df = pd.DataFrame(desc_rows, columns=["Кластер", "Описание", "Критерии"])
                desc_start = start_row + 2 + len(cluster_export)
                desc_df.to_excel(writer, sheet_name="Параметры и кластеры", index=False, startrow=desc_start)
                ws1 = writer.sheets["Параметры и кластеры"]
                for r in range(desc_start + 1, desc_start + 2 + len(desc_df)):
                    for c in range(1, 4):
                        cell = ws1.cell(row=r, column=c)
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Лист 2: цикл жизни
        sheet2_name = "Цикл жизни"
        header_rows = [
            ["Цикл жизни клиента якорного продукта"],
            ["Выбранные кластеры для статистики", ", ".join(lifecycle_clusters) if lifecycle_clusters else "Все"],
            [],
        ]
        header_df = pd.DataFrame(header_rows)
        header_df.to_excel(writer, sheet_name=sheet2_name, index=False, header=False)

        table_start = len(header_rows) + 1
        if lifecycle_table is not None and not lifecycle_table.empty:
            lifecycle_table.to_excel(writer, sheet_name=sheet2_name, index=False, startrow=table_start)

        out_start_row = table_start + (len(lifecycle_table) + 2 if lifecycle_table is not None and not lifecycle_table.empty else 0)

        # Вывод на листе Цикл жизни — структурированный по строкам или одна ячейка
        ws2 = writer.sheets[sheet2_name]
        if lifecycle_output_rows:
            ws2.cell(row=out_start_row + 1, column=1, value="Вывод").font = Font(bold=True)
            current_row = out_start_row + 2
            for row_type, text in lifecycle_output_rows:
                if row_type == "heading":
                    if current_row > out_start_row + 2:
                        ws2.cell(row=current_row, column=1, value="")
                        current_row += 1
                    cell = ws2.cell(row=current_row, column=1, value=text or "")
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    current_row += 1
                elif row_type == "spacer":
                    ws2.cell(row=current_row, column=1, value="")
                    current_row += 1
                else:
                    cell = ws2.cell(row=current_row, column=1, value=text or "")
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    current_row += 1
            ws2.column_dimensions["A"].width = 90
        elif lifecycle_output_text:
            ws2.cell(row=out_start_row + 1, column=1, value="Вывод")
            out_text_cell = ws2.cell(row=out_start_row + 2, column=1, value=lifecycle_output_text)
            out_text_cell.alignment = Alignment(wrap_text=True, vertical="top")
            merge_rows = max(15, min(80, len(lifecycle_output_text) // 60))
            ws2.merge_cells(
                start_row=out_start_row + 2,
                start_column=1,
                end_row=out_start_row + 2 + merge_rows,
                end_column=6,
            )

    buffer.seek(0)
    return buffer.getvalue()


def _placeholder_excel_bytes() -> bytes:
    """Минимальный валидный xlsx для кнопки до первого формирования полного отчёта (после загрузки страницы отчёт подставится автоматически)."""
    buf = io.BytesIO()
    placeholder_df = pd.DataFrame([["Полный отчёт подставится после загрузки страницы. Нажмите «Скачать» ещё раз или обновите страницу."]])
    placeholder_df.to_excel(buf, sheet_name="Инфо", index=False, header=False)
    buf.seek(0)
    return buf.getvalue()


def build_stacked_area(
    df_plot, x_col, value_col, stack_col, title, value_label,
    x_order=None, show_title=True, xaxis_title=None, xaxis_side="bottom",
    margin_override=None,
):
    """Строит стековую диаграмму с областями (stacked area)."""
    if df_plot.empty:
        fig = go.Figure()
        fig.add_annotation(text="Нет данных", xref="paper", yref="paper", x=0.5, y=0.5, showarrow=False)
        fig.update_layout(title=dict(text=title or "", x=0.5, xanchor="center") if show_title and title else {})
        return fig
    x_vals = x_order if x_order is not None else df_plot[x_col].unique().tolist()
    stacks = df_plot[stack_col].unique().tolist()
    fig = go.Figure()
    for s in stacks:
        sub = df_plot[df_plot[stack_col] == s]
        sub = sub.set_index(x_col)[value_col].reindex(x_vals).fillna(0)
        fig.add_trace(
            go.Scatter(
                x=x_vals,
                y=sub.tolist(),
                name=str(s),
                mode="lines",
                fill="tonexty",
                stackgroup="one",
                line=dict(width=0.5),
            )
        )
    margin = margin_override if margin_override is not None else dict(t=60, b=50)
    layout_kw = dict(
        hovermode="x unified",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        margin=margin,
        template="plotly_white",
        yaxis_title=value_label,
    )
    if show_title and title:
        layout_kw["title"] = dict(text=title, x=0.5, xanchor="center")
    if xaxis_title is not None:
        layout_kw["xaxis_title"] = xaxis_title
        layout_kw["xaxis"] = dict(side=xaxis_side)
    else:
        layout_kw["xaxis_title"] = x_col
    fig.update_layout(**layout_kw)
    return fig


# Высота каждого подграфика в объединённой фигуре (пиксели)
COMBINED_CHART_ROW_HEIGHT = 260


def build_combined_two_charts(
    clients_by_period,
    qty_by_period,
    x_col,
    period_labels_short,
    stack_col,
    add_total=False,
    clients_total_values=None,
    qty_total_values=None,
):
    """
    Строит одну фигуру с двумя подграфиками (общая ось X).
    Одинаковые категории — одинаковые цвета в обоих графиках.
    При add_total и 2+ категориях добавляется серия «Итого» (один цвет в легенде).
    """
    x_vals = period_labels_short
    if clients_by_period.empty and qty_by_period.empty:
        fig = go.Figure()
        fig.add_annotation(text="Нет данных", xref="paper", yref="paper", x=0.5, y=0.5, showarrow=False)
        return fig

    stacks_cl = clients_by_period[stack_col].unique().tolist() if not clients_by_period.empty else []
    stacks_q = qty_by_period[stack_col].unique().tolist() if not qty_by_period.empty else []
    all_stacks = list(dict.fromkeys(stacks_cl + stacks_q))
    if add_total:
        all_stacks = ["Итого"] + all_stacks
    palette = px.colors.qualitative.Plotly
    color_map = {s: palette[i % len(palette)] for i, s in enumerate(all_stacks)}

    fig = make_subplots(
        rows=2,
        cols=1,
        shared_xaxes=True,
        vertical_spacing=0.04,
        row_heights=[1, 1],
        subplot_titles=("", ""),
    )

    # Верхний график: клиенты (сначала стек по категориям, затем линия Итого поверх)
    if not clients_by_period.empty:
        for s in stacks_cl:
            sub = clients_by_period[clients_by_period[stack_col] == s]
            sub = sub.set_index(x_col)["clients_count"].reindex(x_vals).fillna(0)
            fig.add_trace(
                go.Scatter(
                    x=x_vals,
                    y=sub.tolist(),
                    name=str(s),
                    mode="lines",
                    fill="tonexty",
                    stackgroup="one",
                    line=dict(width=0.5, color=color_map.get(s, None)),
                ),
                row=1,
                col=1,
            )
        if add_total and clients_total_values is not None:
            fig.add_trace(
                go.Scatter(
                    x=x_vals,
                    y=list(clients_total_values),
                    name="Итого",
                    mode="lines",
                    line=dict(width=1.5, color=color_map.get("Итого", "#636EFA"), dash="dash"),
                ),
                row=1,
                col=1,
            )

    # Нижний график: товар (те же цвета по категориям), легенду не дублируем
    if not qty_by_period.empty:
        for s in stacks_q:
            sub = qty_by_period[qty_by_period[stack_col] == s]
            sub = sub.set_index(x_col)[COL_QUANTITY].reindex(x_vals).fillna(0)
            fig.add_trace(
                go.Scatter(
                    x=x_vals,
                    y=sub.tolist(),
                    name=str(s),
                    mode="lines",
                    fill="tonexty",
                    stackgroup="two",
                    line=dict(width=0.5, color=color_map.get(s, None)),
                    showlegend=False,
                ),
                row=2,
                col=1,
            )
        if add_total and qty_total_values is not None:
            fig.add_trace(
                go.Scatter(
                    x=x_vals,
                    y=list(qty_total_values),
                    name="Итого",
                    mode="lines",
                    line=dict(width=1.5, color=color_map.get("Итого", "#636EFA"), dash="dash"),
                    showlegend=False,
                ),
                row=2,
                col=1,
            )

    total_height = COMBINED_CHART_ROW_HEIGHT * 2
    fig.update_layout(
        height=total_height,
        hovermode="x unified",
        template="plotly_white",
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        margin=dict(t=40, b=40, l=80, r=40),
        hoverlabel=dict(
            namelength=-1,
            font=dict(size=12, color="black"),
            bgcolor="white",
            bordercolor="gray",
        ),
    )
    fig.update_xaxes(title_text="", side="top", row=1, col=1)
    fig.update_xaxes(title_text="", row=2, col=1)
    # Подписи осей Y слева от графика
    fig.update_yaxes(title_text="Количество клиентов", row=1, col=1, side="left")
    fig.update_yaxes(title_text="Количество товара", row=2, col=1, side="left")
    fig.update_xaxes(showspikes=True, spikemode="across+marker", spikecolor="gray", spikethickness=1)
    return fig


# --- Конфигурация страницы (для Streamlit Cloud) ---
st.set_page_config(
    page_title="Цикл жизни клиента в продукте",
    page_icon="🔄",
    layout="wide",
)

# --- Заголовок (отдельно от контента) ---
st.title("🔄 Цикл жизни клиента в продукте")
st.divider()

# --- Две колонки: инструкция слева, шаблон и загрузчик справа ---
col_instruction, col_template = st.columns([1, 1])

with col_instruction:
    st.subheader("Инструкция к загрузке 1 документа")
    st.markdown("""
    1. Зайдите в Qlik, раздел «Анализ чеков», лист «Конструктор».
    2. Отберите анализируемые продукты/категорию в одном из разрезов Группа1 / Группа2 / Группа3 / Группа4.
    3. Отберите анализируемый период и разрез (год–месяц или год–неделя).
    4. Выведите отчёт по шаблону справа.
    5. Скачайте документ в Qlik и загрузите в ячейку справа.
    """)
    st.markdown("---")
    st.subheader("Инструкция к загрузке 2 документа")
    st.markdown("""
    1. Перейдите на лист «Продажи» и отберите клиентов анализируемого периода и продукта/категории.
    2. Выберите альтернативные категории и выведите отчёт по шаблону справа на листе «Конструктор».
    3. Скачайте документ в Qlik и загрузите в ячейку справа.
    """)

with col_template:
    st.subheader("📋 Шаблон загрузки данных из Qlik")
    try:
        st.image("qlik_template_categories.png", use_container_width=True)
    except FileNotFoundError:
        st.warning("Изображение шаблона не найдено. Убедитесь, что файл `qlik_template_categories.png` в корне проекта.")
    st.markdown("---")
    st.caption("Документ 1 — выгрузка из листа «Конструктор»:")
    uploaded_file_1 = st.file_uploader(
        "Документ 1",
        type=["xlsx", "xls"],
        key="qlik_upload_1",
        label_visibility="collapsed",
    )
    st.caption("Документ 2 — выгрузка по альтернативным категориям:")
    uploaded_file_2 = st.file_uploader(
        "Документ 2",
        type=["xlsx", "xls"],
        key="qlik_upload_2",
        label_visibility="collapsed",
    )

# --- Блок отображения расчётов (графики) — показывается после загрузки обоих документов ---
if uploaded_file_1 and uploaded_file_2:
    st.divider()
    try:
        df1 = load_and_normalize(uploaded_file_1)
        df2 = load_and_normalize(uploaded_file_2)
    except Exception as e:
        st.error(f"Ошибка чтения файлов: {e}")
        df1 = df2 = None

    if df1 is not None and df2 is not None and not df1.empty and not df2.empty:
        had_excel_bytes = "excel_report_bytes" in st.session_state
        categories_from_doc1 = sorted(df1[COL_CATEGORY].dropna().unique().tolist())
        category_label = ", ".join(categories_from_doc1) if categories_from_doc1 else "—"
        st.markdown(f"### Якорный продукт когорт: :violet[{category_label}]")

        df, period_order, rank_to_period, _ = merge_and_prepare(df1, df2)
        period_labels_short = [
            format_period_short(row[COL_PERIOD_MAIN], row[COL_PERIOD_SUB])
            for _, row in period_order[[COL_PERIOD_MAIN, COL_PERIOD_SUB]].iterrows()
        ]
        period_rank_to_short = dict(zip(period_order["period_rank"], period_labels_short))
        categories_from_doc2 = sorted(df2[COL_CATEGORY].dropna().unique().tolist())
        categories_from_doc1_set = set(categories_from_doc1)
        # Все категории (док 1 + док 2) — для расчёта «прочих» в блоке цикла жизни
        all_categories = categories_from_doc1 + [c for c in categories_from_doc2 if c not in categories_from_doc1_set]
        # Список только из документа 2 — якорный продукт (документ 1) нельзя выбирать как анализируемый
        analyzable_product_options = [c for c in categories_from_doc2 if c not in categories_from_doc1_set]
        # Когорта клиента = первая покупка якорного (док 1). Размер когорты = число клиентов с первой покупкой в этом периоде.
        df1_for_cohort = df1.copy()
        df1_for_cohort[COL_PERIOD_MAIN] = df1_for_cohort[COL_PERIOD_MAIN].astype(str).str.strip()
        df1_for_cohort[COL_PERIOD_SUB] = df1_for_cohort[COL_PERIOD_SUB].astype(str).str.strip()
        df1_for_cohort = df1_for_cohort.merge(
            period_order[[COL_PERIOD_MAIN, COL_PERIOD_SUB, "period_rank"]],
            on=[COL_PERIOD_MAIN, COL_PERIOD_SUB],
            how="left",
        )
        df1_for_cohort["_client_norm"] = _norm_client_id(df1_for_cohort[COL_CLIENT])
        client_first_rank = df1_for_cohort.groupby("_client_norm")["period_rank"].min()
        # Когорты с подписью вида "2025/01 (N клиентов)" — N = размер когорты (первая покупка в этом периоде)
        cohort_options = []
        for r in sorted(rank_to_period.index):
            row = rank_to_period.loc[r]
            pm, ps = str(row[COL_PERIOD_MAIN]).strip(), str(row[COL_PERIOD_SUB]).strip()
            short = period_labels_short[r] if r < len(period_labels_short) else f"{pm} {ps}"
            n_clients = (client_first_rank == r).sum()
            label = f"{short} ({n_clients} клиентов)"
            cohort_options.append((r, label))
        cohort_labels = [lb for _, lb in cohort_options]
        cohort_ranks = {lb: r for r, lb in cohort_options}

        # Подготовка данных с period_rank для блоков ниже (кластеры, цикл жизни, продажи)
        df1_norm = df1.copy()
        df1_norm[COL_PERIOD_MAIN] = df1_norm[COL_PERIOD_MAIN].astype(str).str.strip()
        df1_norm[COL_PERIOD_SUB] = df1_norm[COL_PERIOD_SUB].astype(str).str.strip()
        df2_norm = df2.copy()
        df2_norm[COL_PERIOD_MAIN] = df2_norm[COL_PERIOD_MAIN].astype(str).str.strip()
        df2_norm[COL_PERIOD_SUB] = df2_norm[COL_PERIOD_SUB].astype(str).str.strip()
        df1_with_period = df1_norm.merge(
            period_order[[COL_PERIOD_MAIN, COL_PERIOD_SUB, "period_rank"]],
            on=[COL_PERIOD_MAIN, COL_PERIOD_SUB],
            how="left",
        )
        df1_with_period["period_label_short"] = df1_with_period["period_rank"].map(period_rank_to_short)
        df2_with_period = df2_norm.merge(
            period_order[[COL_PERIOD_MAIN, COL_PERIOD_SUB, "period_rank"]],
            on=[COL_PERIOD_MAIN, COL_PERIOD_SUB],
            how="left",
        )
        df2_with_period["period_label_short"] = df2_with_period["period_rank"].map(period_rank_to_short)
        df2_with_period["_client_norm"] = _norm_client_id(df2_with_period[COL_CLIENT])
        df1_with_period["_client_norm"] = _norm_client_id(df1_with_period[COL_CLIENT])

        # Определение типа периода по данным (недели или месяцы) — используется во всех блоках
        period_sub_str = period_order[COL_PERIOD_SUB].astype(str).str.lower()
        is_months = period_sub_str.str.contains(r"янв|фев|мар|апр|май|июн|июл|авг|сен|окт|ноя|дек", regex=True).any()
        period_word = "месяцев" if is_months else "недель"

        # --- Настройка параметров отчёта (применяется ко всем блокам) ---
        st.divider()
        st.subheader("Настройка параметров отчёта")
        # Подтягиваем кнопку скачивания вверх, чтобы низ совпадал с селектом «По когорту»
        st.markdown(
            """
            <style>
            div[data-testid="column"]:has(iframe[height="46"]) {
                margin-top: 22px !important;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )
        # Первый ряд: [С когорты | По когорту] | Анализируемый продукт (уже) | Недель/месяцев
        r1_c1, r1_c2, r1_c3 = st.columns([1.2, 0.7, 1])
        with r1_c1:
            sub_left, sub_right = st.columns(2)
            with sub_left:
                cohort_start_global = st.selectbox(
                    "С когорты",
                    options=cohort_labels,
                    index=0,
                    key="report_cohort_start",
                )
            with sub_right:
                cohort_end_global = st.selectbox(
                    "По когорту",
                    options=cohort_labels,
                    index=0,
                    key="report_cohort_end",
                )
        with r1_c2:
            selected_categories_global = st.multiselect(
                "Анализируемый продукт",
                options=analyzable_product_options,
                default=[],
                key="report_categories",
                help="Категории из документа 2 для кластеризации, цикла жизни и расчёта продаж на объём якорного. Якорный продукт (документ 1) в списке не отображается.",
            )
        with r1_c3:
            k_periods_global = st.number_input(
                "Недель/месяцев с покупки якорного (включая период когорты)",
                min_value=1,
                value=5,
                step=1,
                key="report_k_periods",
            )
        # Второй ряд: пусто | пусто | Кнопка
        r2_c1, r2_c2, r2_c3 = st.columns([1.2, 0.7, 1])
        with r2_c1:
            pass
        with r2_c2:
            pass
        with r2_c3:
            excel_data = st.session_state.get("excel_report_bytes") or _placeholder_excel_bytes()
            report_filename = st.session_state.get("excel_report_filename", "CLF_report.xlsx")
            create_excel_download_button(
                excel_data,
                report_filename,
                "Скачать полный отчёт в Excel",
                "download_full_report",
            )

        idx_start_c = cohort_labels.index(cohort_start_global)
        idx_end_c = cohort_labels.index(cohort_end_global)
        if idx_start_c <= idx_end_c:
            cohorts_to_use_c = cohort_labels[idx_start_c : idx_end_c + 1]
        else:
            cohorts_to_use_c = cohort_labels[idx_end_c : idx_start_c + 1]
        selected_categories_cluster = selected_categories_global
        k_periods_cluster = k_periods_global

        # --- Блок «Кластерный анализ» (без своих настроек — используются из «Настройка параметров отчёта») ---
        st.divider()
        st.subheader("Кластерный анализ")
        st.caption("Сегментация клиентов по объёму покупок и регулярности покупок выбранного продукта в первые K периодов после когорты.")

        if not cohorts_to_use_c:
            st.caption("Выберите хотя бы одну когорту в настройках выше.")
        elif not selected_categories_cluster:
            st.warning("Выберите анализируемый продукт для кластеризации.")
        else:
            # Клиенты, чья первая покупка якорного (док 1) попадает в выбранный диапазон когорт
            selected_ranks_c = {cohort_ranks[lb] for lb in cohorts_to_use_c}
            cohort_clients_c = set(client_first_rank.index[client_first_rank.isin(selected_ranks_c)].tolist())

            if not cohort_clients_c:
                st.info("В выбранных когортах нет клиентов (по первой покупке якорного продукта).")
            else:
                max_rank = int(period_order["period_rank"].max())
                k_int = int(k_periods_cluster)
                cohorts_short_data_c = [lb for lb in cohorts_to_use_c if cohort_ranks[lb] > max_rank - k_int + 1]
                if cohorts_short_data_c:
                    short_labels_c = [
                        format_period_short(rank_to_period.loc[cohort_ranks[lb], COL_PERIOD_MAIN], rank_to_period.loc[cohort_ranks[lb], COL_PERIOD_SUB])
                        for lb in cohorts_short_data_c
                    ]
                    st.warning(f"Увеличьте период данных или уменьшите кол-во недель/месяцев с покупки якорного продукта для корректного расчёта когорт ({', '.join(short_labels_c)})")

                # Для каждого клиента — его период когорты (min period_rank по документу 1)
                df1_cr = df1_with_period.copy()
                df1_cr["_client_norm"] = _norm_client_id(df1_cr[COL_CLIENT])
                df1_cr = df1_cr[df1_cr["_client_norm"].isin(cohort_clients_c)]
                client_cohort_rank = df1_cr.groupby("_client_norm")["period_rank"].min()

                # Сколько периодов доступно для наблюдения (для поздних когорт окно может упираться в конец данных)
                available_periods = (max_rank - client_cohort_rank + 1).clip(lower=0, upper=k_int).astype(int)

                # Собираем покупки анализируемого продукта из документа 1 и/или документа 2
                selected_in_doc1_c = [c for c in selected_categories_cluster if c in categories_from_doc1_set]
                selected_in_doc2_c = [c for c in selected_categories_cluster if c in set(categories_from_doc2)]

                def _filter_to_dynamic_window(df_src: pd.DataFrame) -> pd.DataFrame:
                    """Фильтрует строки когорты в окне [cohort_rank, cohort_rank + K)."""
                    tmp = df_src.copy()
                    tmp["_client_norm"] = _norm_client_id(tmp[COL_CLIENT])
                    tmp = tmp[tmp["_client_norm"].isin(cohort_clients_c)]
                    r0 = tmp["_client_norm"].map(client_cohort_rank)
                    delta = tmp["period_rank"] - r0
                    mask = delta.notna() & tmp["period_rank"].notna() & (delta >= 0) & (delta < k_int)
                    return tmp.loc[mask, ["_client_norm", "period_rank", COL_QUANTITY]]

                parts_p = []
                if selected_in_doc1_c:
                    parts_p.append(
                        _filter_to_dynamic_window(
                            df1_with_period[df1_with_period[COL_CATEGORY].isin(selected_in_doc1_c)]
                        )
                    )
                if selected_in_doc2_c:
                    parts_p.append(
                        _filter_to_dynamic_window(
                            df2_with_period[df2_with_period[COL_CATEGORY].isin(selected_in_doc2_c)]
                        )
                    )

                if parts_p:
                    df_p = pd.concat(parts_p, ignore_index=True)
                else:
                    df_p = pd.DataFrame(columns=["_client_norm", "period_rank", COL_QUANTITY])

                # Метрики по клиенту: объём и регулярность (доля периодов с покупкой)
                per_client = pd.DataFrame({"client_id": sorted(cohort_clients_c)})
                per_client["cohort_rank"] = per_client["client_id"].map(client_cohort_rank).astype("float")
                per_client["available_periods"] = per_client["client_id"].map(available_periods).fillna(k_int).astype(int)

                if not df_p.empty:
                    agg = (
                        df_p.groupby("_client_norm")
                        .agg(
                            volume=(COL_QUANTITY, "sum"),
                            active_periods=("period_rank", "nunique"),
                        )
                        .reset_index()
                        .rename(columns={"_client_norm": "client_id"})
                    )
                    per_client = per_client.merge(agg, on="client_id", how="left")
                per_client["volume"] = per_client["volume"].fillna(0).astype(int)
                per_client["active_periods"] = per_client["active_periods"].fillna(0).astype(int)
                denom = per_client["available_periods"].replace(0, 1)
                per_client["regularity"] = (per_client["active_periods"] / denom).clip(0, 1).astype(float)

                # Назначение одному из 8 поведенческих кластеров по объёму и регулярности (правила по перцентилям)
                per_client["cluster"] = "Не покупали"
                df_fit = per_client[per_client["volume"] > 0].copy()
                v33_val = v67_val = 0.0
                if not df_fit.empty:
                    v33_val = float(df_fit["volume"].quantile(1 / 3))
                    v67_val = float(df_fit["volume"].quantile(2 / 3))
                    v33 = v33_val
                    v67 = v67_val
                    r33 = 1 / 3
                    r67 = 2 / 3

                    def _assign_cluster(row):
                        v, r = row["volume"], row["regularity"]
                        if v >= v67:
                            if r >= r67:
                                return "Активные (VIP)"
                            if r >= r33:
                                return "Регулярные с высоким объёмом"
                            return "Разовая крупная покупка"
                        if v >= v33:
                            if r >= r67:
                                return "Средняя активность"
                            if r >= r33:
                                return "Средняя активность"
                            return "Крупные нерегулярные"
                        if r >= r67:
                            return "Периодические (малый объём)"
                        if r >= r33:
                            return "Низкая активность"
                        return "Разовая покупка"

                    df_fit["cluster"] = df_fit.apply(_assign_cluster, axis=1)
                    per_client = per_client.merge(df_fit[["client_id", "cluster"]], on="client_id", how="left", suffixes=("", "_fit"))
                    per_client["cluster"] = per_client["cluster_fit"].fillna(per_client["cluster"])
                    per_client = per_client.drop(columns=["cluster_fit"], errors="ignore")

                total_clients = len(per_client)
                k_int_cluster = int(k_periods_cluster)
                period_unit = "месяц" if is_months else "неделю"
                summary = (
                    per_client.groupby("cluster", dropna=False)
                    .agg(
                        clients=("client_id", "count"),
                        pct=("client_id", lambda s: 100.0 * len(s) / total_clients if total_clients else 0.0),
                        total_volume=("volume", "sum"),
                        active_periods_sum=("active_periods", "sum"),
                        avg_regularity=("regularity", "mean"),
                    )
                    .reset_index()
                )
                # Средний объём на клиента в неделю/месяц покупки (делим на число периодов с покупкой, а не на K)
                summary["avg_client_per_period"] = (
                    (summary["total_volume"] / summary["active_periods_sum"].replace(0, 1)).round(2)
                )
                # Всегда 8 поведенческих кластеров + Не покупали; недостающие — 0
                for c in CLUSTER_8_ORDER:
                    if c not in summary["cluster"].values:
                        summary = pd.concat(
                            [
                                summary,
                                pd.DataFrame(
                                    [{
                                        "cluster": c,
                                        "clients": 0,
                                        "pct": 0.0,
                                        "total_volume": 0,
                                        "active_periods_sum": 0,
                                        "avg_regularity": 0.0,
                                        "avg_client_per_period": 0.0,
                                    }]
                                ),
                            ],
                            ignore_index=True,
                        )
                if "Не покупали" not in summary["cluster"].values:
                    summary = pd.concat(
                        [
                            summary,
                                pd.DataFrame(
                                    [{
                                        "cluster": "Не покупали",
                                        "clients": 0,
                                        "pct": 0.0,
                                        "total_volume": 0,
                                        "active_periods_sum": 0,
                                        "avg_regularity": 0.0,
                                        "avg_client_per_period": 0.0,
                                    }]
                                ),
                        ],
                        ignore_index=True,
                    )
                # Порядок: по объёму за период от большего к меньшему, нулевые кластеры в конце
                summary = summary.sort_values("total_volume", ascending=False).reset_index(drop=True)

                total_volume_all = per_client["volume"].sum()
                total_active_periods_all = per_client["active_periods"].sum()
                avg_client_per_period_all = (
                    total_volume_all / total_active_periods_all if total_active_periods_all else 0
                )
                avg_regularity_all = per_client["regularity"].mean()
                row_итого = pd.DataFrame(
                    [{
                        "cluster": "Итого",
                        "clients": total_clients,
                        "pct": 100.0,
                        "total_volume": int(total_volume_all),
                        "avg_client_per_period": round(avg_client_per_period_all, 2),
                        "avg_regularity": round(avg_regularity_all, 3),
                    }]
                )
                summary = pd.concat([row_итого, summary], ignore_index=True)
                summary["total_volume"] = summary["total_volume"].astype(int)
                # Доля объёма по кластеру от общего объёма за период (%)
                denom_vol = total_volume_all if total_volume_all else 1
                summary["pct_volume"] = (100.0 * summary["total_volume"] / denom_vol).round(1)
                summary["pct_volume_fmt"] = summary["pct_volume"].astype(str) + "%"

                st.session_state["report_cluster_summary"] = summary.copy()

                col_cluster = "Кластер"
                col_volume = "Объём анализируемого продукта за период"
                col_pct_volume = "% объёма"
                col_avg_client = f"Средний объём на клиента в {period_unit} покупки"
                period_word_plural = "месяцев" if is_months else "недель"
                days_per_period = 30 if is_months else 7
                summary["pct_fmt"] = summary["pct"].round(1).astype(str) + "%"

                def _criteria_text(name: str, v33: float, v67: float, k: int, is_m: bool) -> str:
                    v33s = f"{v33:.0f}" if v33 == int(v33) else f"{v33:.1f}"
                    v67s = f"{v67:.0f}" if v67 == int(v67) else f"{v67:.1f}"
                    pw = "месяцев" if is_m else "недель"
                    dp = 30 if is_m else 7
                    n33 = max(1, round(1 / 3 * k))
                    n67 = max(1, round(2 / 3 * k))

                    def _days(ratio: float) -> str:
                        if ratio <= 0:
                            return "—"
                        d = round(dp / ratio)
                        return f"{max(1, int(d))} дн."

                    if name == "Активные (VIP)":
                        return f"Объём ≥ {v67s} ед. (верхняя треть) за весь период. Присутствуют не реже {n67} {pw} из {k} (67%). Приходят не реже чем каждые {_days(2/3)}."
                    if name == "Регулярные с высоким объёмом":
                        return f"Объём ≥ {v67s} ед. за весь период. Присутствуют в {n33}–{n67} {pw} из {k} (33–67%). Приходят в среднем каждые {_days(0.5)}–{_days(1/3)}."
                    if name == "Разовая крупная покупка":
                        return f"Объём ≥ {v67s} ед. за весь период. Присутствуют реже {n33} {pw} из {k} (<33%). Приходят реже чем каждые {_days(0.33)}."
                    if name == "Средняя активность":
                        return f"Объём {v33s}–{v67s} ед. (средняя треть) за весь период. Присутствуют не реже {n33} {pw} из {k} (33%). Приходят не реже чем каждые {_days(1/3)}."
                    if name == "Крупные нерегулярные":
                        return f"Объём {v33s}–{v67s} ед. за весь период. Присутствуют реже {n33} {pw} из {k} (<33%). Приходят реже чем каждые {_days(0.33)}."
                    if name == "Периодические (малый объём)":
                        return f"Объём < {v33s} ед. (нижняя треть) за весь период. Присутствуют не реже {n67} {pw} из {k} (67%). Приходят не реже чем каждые {_days(2/3)}."
                    if name == "Низкая активность":
                        return f"Объём < {v33s} ед. за весь период. Присутствуют в {n33}–{n67} {pw} из {k} (33–67%). Приходят в среднем каждые {_days(0.5)}–{_days(1/3)}."
                    if name == "Разовая покупка":
                        return f"Объём < {v33s} ед. за весь период. Присутствуют реже {n33} {pw} из {k} (<33%). Приходят реже чем каждые {_days(0.33)} или одна покупка."
                    if name == "Не покупали":
                        return "Нет покупок анализируемого продукта в выбранном окне."
                    return ""

                report_cluster_comments = {}
                for _, r in summary.iterrows():
                    cn = r["cluster"]
                    if cn == "Итого":
                        continue
                    desc_text = CLUSTER_8_DESCRIPTIONS.get(cn, "")
                    crit_text = _criteria_text(cn, v33_val, v67_val, k_int_cluster, is_months)
                    report_cluster_comments[cn] = desc_text + ("\n\nКритерии: " + crit_text if crit_text else "")
                st.session_state["report_cluster_comments"] = report_cluster_comments

                cluster_names_list = summary["cluster"].tolist()
                cluster_options = [c for c in cluster_names_list if c != "Итого"]
                cluster_options_display = [_cluster_display_name(n) for n in cluster_options]
                cluster_display_to_full = {_cluster_display_name(n): n for n in cluster_names_list}
                cluster_full_to_display = {n: _cluster_display_name(n) for n in cluster_names_list}

                desc = CLUSTER_8_DESCRIPTIONS
                col_presence = "Присутствие"
                col_regularity_2 = "Регулярность"

                def _escape_html(s: str) -> str:
                    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

                num_columns = 8
                rows_html = []
                for row_idx, r in summary.iterrows():
                    cluster_name = r["cluster"]
                    crit = _criteria_text(cluster_name, v33_val, v67_val, k_int_cluster, is_months)
                    if cluster_name == "Итого":
                        cell_cluster = "<strong>Итого</strong>"
                    else:
                        display_name = _cluster_display_name(cluster_name)
                        desc_text = desc.get(cluster_name, "") or ""
                        cell_cluster = (
                            f'<details class="cluster-details-wrap">'
                            f'<summary class="cluster-summary"><span class="cluster-arrow">▶</span> {_escape_html(display_name)}</summary>'
                            f"</details>"
                        )
                        details_row_content = (
                            f'<div class="cluster-details">'
                            f'<div class="cluster-details-inner">'
                            f'<div class="cluster-detail-block"><strong>Описание:</strong> {_escape_html(desc_text)}</div>'
                            f'<div class="cluster-detail-block"><strong>Критерии:</strong> {_escape_html(crit)}</div>'
                            f"</div></div>"
                        )
                    pct_val = r["pct_fmt"]
                    avg_r = r["avg_regularity"] if pd.notna(r["avg_regularity"]) else 0
                    x_per = round(avg_r * k_int_cluster, 1)
                    y_pct = round(avg_r * 100, 1)
                    line1 = f"{x_per} {period_word_plural} из {k_int_cluster} ({y_pct}%)"
                    if avg_r > 0.001:
                        z_days = max(1, round(days_per_period / avg_r))
                        window_days = k_int_cluster * days_per_period
                        suffix = " (вероятно реже)" if z_days >= window_days else ""
                        line2 = f"В среднем каждые {z_days} дн.{suffix}"
                    else:
                        line2 = "Приходят редко или одна покупка"
                    rows_html.append(
                        f"<tr><td>{cell_cluster}</td>"
                        f"<td>{int(r['clients'])}</td><td>{pct_val}</td>"
                        f"<td>{int(r['total_volume'])}</td><td>{r['pct_volume_fmt']}</td><td>{r['avg_client_per_period']:.2f}</td>"
                        f"<td>{line1}</td><td>{line2}</td></tr>"
                    )
                    if cluster_name != "Итого":
                        rows_html.append(
                            f'<tr class="cluster-details-row" style="display:none;">'
                            f'<td colspan="{num_columns}">{details_row_content}</td></tr>'
                        )
                thead = (
                    f"<thead><tr>"
                    f"<th>{col_cluster}</th>"
                    f"<th>Клиентов</th><th>% клиентов</th><th>{col_volume}</th><th>{col_pct_volume}</th><th>{col_avg_client}</th>"
                    f"<th>{col_presence}</th><th>{col_regularity_2}</th>"
                    f"</tr></thead>"
                )
                tbody = "<tbody>" + "".join(rows_html) + "</tbody>"
                cluster_table_html = (
                    "<!DOCTYPE html><html><head><meta charset='utf-8'>"
                    "<style>\n"
                    "html, body { overflow-x: hidden; max-width: 100%; box-sizing: border-box; }\n"
                    "body { font-family: 'Source Sans 3', 'Source Sans Pro', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; font-size: 1rem; margin: 0; padding: 0.5rem; }\n"
                    ".cluster-table-shell { margin: 0.5rem 0; background: #dee2e6; padding: 8px; border-radius: 8px; overflow: hidden; }\n"
                    ".cluster-table-wrap { min-height: 400px; max-height: 85vh; overflow-y: auto; overflow-x: hidden; width: 100%; max-width: 100%; }\n"
                    ".cluster-table { width: 100%; max-width: 100%; table-layout: fixed; border-collapse: collapse; font-size: 0.9375rem; "
                    "border: 1px solid #adb5bd; border-radius: 0; box-shadow: 0 2px 6px rgba(0,0,0,0.06); background: #e2e6ea; }\n"
                    ".cluster-table thead th { position: sticky; top: 0; z-index: 100; box-sizing: border-box; "
                    "background: #343a40; color: #fff; font-weight: 600; padding: 6px 10px; text-align: left; "
                    "font-size: 0.9375rem; box-shadow: 0 2px 2px rgba(0,0,0,0.2); line-height: 1.3; "
                    "word-wrap: break-word; overflow-wrap: break-word; border: 1px solid #adb5bd; }\n"
                    ".cluster-table thead th:nth-child(1) { width: 16%; }\n"
                    ".cluster-table thead th:nth-child(2) { width: 8%; }\n"
                    ".cluster-table thead th:nth-child(3) { width: 8%; }\n"
                    ".cluster-table thead th:nth-child(4) { width: 10%; }\n"
                    ".cluster-table thead th:nth-child(5) { width: 8%; }\n"
                    ".cluster-table thead th:nth-child(6) { width: 12%; }\n"
                    ".cluster-table thead th:nth-child(7) { width: 14%; }\n"
                    ".cluster-table thead th:nth-child(8) { width: 12%; }\n"
                    ".cluster-table thead th:nth-child(9) { width: 12%; }\n"
                    ".cluster-table thead th:nth-child(2), .cluster-table thead th:nth-child(3), .cluster-table thead th:nth-child(4), "
                    ".cluster-table thead th:nth-child(5), .cluster-table thead th:nth-child(6) { text-align: center; }\n"
                    ".cluster-table td { padding: 5px 10px; border: 1px solid #adb5bd; background: #e2e6ea; color: #212529; vertical-align: middle; font-size: 0.9375rem; line-height: 1.35; "
                    "word-wrap: break-word; overflow-wrap: break-word; box-sizing: border-box; }\n"
                    ".cluster-table td:nth-child(1) { font-weight: 500; }\n"
                    ".cluster-table td:nth-child(2), .cluster-table td:nth-child(3), .cluster-table td:nth-child(4), "
                    ".cluster-table td:nth-child(5), .cluster-table td:nth-child(6) { text-align: center; }\n"
                    ".cluster-details-wrap summary { list-style: none; cursor: pointer; font-size: 0.9375rem; line-height: 1.3; }\n"
                    ".cluster-details-wrap summary::-webkit-details-marker { display: none; }\n"
                    ".cluster-arrow { display: inline-block; margin-right: 4px; font-size: 0.7rem; color: #495057; }\n"
                    ".cluster-details-wrap[open] .cluster-arrow { transform: rotate(90deg); }\n"
                    ".cluster-details-row td { padding: 0; border: 1px solid #adb5bd; background: #e2e6ea; vertical-align: top; }\n"
                    ".cluster-details { padding: 12px 16px; margin: 8px 10px; background: #fff; color: #212529; "
                    "border: 1px solid #adb5bd; border-left: 4px solid #495057; border-radius: 6px; font-size: 0.9rem; line-height: 1.5; "
                    "box-shadow: 0 2px 8px rgba(0,0,0,0.08); }\n"
                    ".cluster-details-inner { display: grid; grid-template-columns: 1fr 1fr; gap: 1rem 1.5rem; max-width: 100%; }\n"
                    ".cluster-detail-block { min-width: 0; }\n"
                    ".cluster-detail-block strong { font-weight: 600; font-style: italic; display: block; margin-bottom: 2px; }\n"
                    ".cluster-detail-block:first-child strong { color: #5c2d91; }\n"
                    ".cluster-detail-block:last-child strong { color: #e85d04; }\n"
                    "@media (max-width: 640px) { .cluster-details-inner { grid-template-columns: 1fr; } }\n"
                    ".cluster-table tbody tr:hover td { background-color: #e9ecef; }\n"
                    ".cluster-table tbody tr:first-child td { background: rgba(128, 0, 128, 0.4) !important; color: #fff !important; font-weight: bold; }\n"
                    ".cluster-table tbody tr:first-child:hover td { background: rgba(128, 0, 128, 0.5) !important; }\n"
                    ".cluster-table tbody tr:first-child .cluster-arrow { color: rgba(255,255,255,0.9); }\n"
                    "</style></head><body>"
                    f'<div class="cluster-table-shell"><div class="cluster-table-wrap"><table class="cluster-table">{thead}{tbody}</table></div></div>'
                    "<script>"
                    "document.querySelectorAll('.cluster-details-wrap').forEach(function(d){"
                    "  d.addEventListener('toggle', function(){"
                    "    var tr = this.closest('tr');"
                    "    if(!tr) return;"
                    "    var next = tr.nextElementSibling;"
                    "    if(next && next.classList.contains('cluster-details-row')){"
                    "      next.style.display = this.open ? 'table-row' : 'none';"
                    "    }"
                    "  });"
                    "});"
                    "</script>"
                    "</body></html>"
                )
                components.html(cluster_table_html, height=min(520, 180 + len(rows_html) * 32), scrolling=True)

                # Выбор кластеров для копирования кодов клиента (прямо под таблицей)
                st.markdown("<div style='margin-top: 0.25rem;'></div>", unsafe_allow_html=True)
                col_clusters_sel, col_copy_btn = st.columns([1, 1])
                with col_clusters_sel:
                    selected_clusters_for_copy = st.multiselect(
                        "Выбор кластеров для копирования кодов клиента",
                        options=cluster_options_display,
                        default=[],
                        key="cluster_copy_multiselect",
                    )
                with col_copy_btn:
                    selected_full_names = [cluster_display_to_full[s] for s in selected_clusters_for_copy if s in cluster_display_to_full]
                    ids_for_copy = per_client[per_client["cluster"].isin(selected_full_names)]["client_id"].tolist()
                    copy_data_str = "\n".join(str(c) for c in ids_for_copy)
                    n_copy = len(ids_for_copy)
                    copy_label = f"📋 Копировать коды ({n_copy})" if n_copy else "📋 Копировать коды (0)"
                    create_copy_button(copy_data_str, copy_label, "copy_cluster_codes")

        # --- Блок «Цикл жизни клиента якорного продукта» (настройки — из «Настройка параметров отчёта», только выбор кластеров) ---
        st.divider()
        st.subheader("Цикл жизни клиента якорного продукта")

        cohort_start_lc = cohort_start_global
        cohort_end_lc = cohort_end_global
        selected_categories_lifecycle = selected_categories_global
        k_periods_lifecycle = k_periods_global

        idx_start_lc = cohort_labels.index(cohort_start_lc)
        idx_end_lc = cohort_labels.index(cohort_end_lc)
        if idx_start_lc <= idx_end_lc:
            cohorts_to_use_lc = cohort_labels[idx_start_lc : idx_end_lc + 1]
        else:
            cohorts_to_use_lc = cohort_labels[idx_end_lc : idx_start_lc + 1]

        st.caption("Отбор кластеров для статистики по неделям.")
        cluster_options_only = [_cluster_display_name(c) for c in CLUSTER_8_ORDER] + ["Не покупали"]
        cluster_options_with_all = ["Все кластеры"] + cluster_options_only
        if "lifecycle_clusters_multiselect" not in st.session_state:
            st.session_state["lifecycle_clusters_multiselect"] = ["Все кластеры"]
        raw_selection = st.multiselect(
            "Кластеры",
            options=cluster_options_with_all,
            key="lifecycle_clusters_multiselect",
            label_visibility="collapsed",
        )
        # Авто-логика: если выбраны и «Все кластеры», и другие — оставляем только другие
        if "Все кластеры" in raw_selection and len(raw_selection) > 1:
            selected_clusters_lifecycle = [x for x in raw_selection if x != "Все кластеры"]
        else:
            selected_clusters_lifecycle = raw_selection if raw_selection else ["Все кластеры"]
        if raw_selection != selected_clusters_lifecycle:
            st.session_state["lifecycle_clusters_multiselect"] = selected_clusters_lifecycle
            st.rerun()

        if not cohorts_to_use_lc:
            st.caption("Выберите хотя бы одну когорту.")
        elif not selected_categories_lifecycle:
            st.warning("Выберите хотя бы один анализируемый продукт.")
        else:
            # Клиенты, чья первая покупка якорного (док 1) попадает в выбранный диапазон когорт
            selected_ranks_lc = {cohort_ranks[lb] for lb in cohorts_to_use_lc}
            cohort_clients_lc = set(client_first_rank.index[client_first_rank.isin(selected_ranks_lc)].tolist())

            if not cohort_clients_lc:
                st.info("В выбранных когортах нет клиентов (по первой покупке якорного продукта).")
            else:
                df1_cr_lc = df1_with_period.copy()
                df1_cr_lc["_client_norm"] = _norm_client_id(df1_cr_lc[COL_CLIENT])
                df1_cr_lc = df1_cr_lc[df1_cr_lc["_client_norm"].isin(cohort_clients_lc)]
                client_cohort_rank_lc = df1_cr_lc.groupby("_client_norm")["period_rank"].min()

                k_int_lc = int(k_periods_lifecycle)
                n_anchor_lc = 100  # фиксированное число для расчёта и фразы «При продаже … будет продано …»
                client_cohort_rank_dict_lc = client_cohort_rank_lc.to_dict()

                max_rank_lc = int(period_order["period_rank"].max())
                cohorts_short_data_lc = [lb for lb in cohorts_to_use_lc if cohort_ranks[lb] > max_rank_lc - k_int_lc + 1]
                if cohorts_short_data_lc:
                    short_labels_lc = [
                        format_period_short(rank_to_period.loc[cohort_ranks[lb], COL_PERIOD_MAIN], rank_to_period.loc[cohort_ranks[lb], COL_PERIOD_SUB])
                        for lb in cohorts_short_data_lc
                    ]
                    st.warning(f"Увеличьте период данных или уменьшите кол-во недель/месяцев с покупки якорного продукта для корректного расчёта когорт ({', '.join(short_labels_lc)})")

                def _in_window_lc(row):
                    c = row.get("_client_norm")
                    r0 = client_cohort_rank_dict_lc.get(c)
                    if r0 is None or pd.isna(r0):
                        return False
                    pr = row.get("period_rank")
                    if pd.isna(pr):
                        return False
                    return r0 <= pr < r0 + k_int_lc

                # Категории и покупки по когорте — нужны для кластеров и для расчёта продаж по кластерам
                anchor_cats = set(categories_from_doc1)
                analyzable_list = list(selected_categories_lifecycle)
                other_cats = set(all_categories) - anchor_cats - set(analyzable_list)

                df1_lc = df1_with_period[df1_with_period["_client_norm"].isin(cohort_clients_lc)][
                    ["_client_norm", "period_rank", COL_CATEGORY, COL_QUANTITY]
                ].copy()
                df2_lc = df2_with_period[df2_with_period["_client_norm"].isin(cohort_clients_lc)][
                    ["_client_norm", "period_rank", COL_CATEGORY, COL_QUANTITY]
                ].copy()
                df_purchases_lc = pd.concat([df1_lc.rename(columns={"_client_norm": "client_id"}), df2_lc.rename(columns={"_client_norm": "client_id"})], ignore_index=True)

                def _to_set(x):
                    return x if isinstance(x, set) else set()

                client_period_cats = (
                    df_purchases_lc.groupby(["client_id", "period_rank"])[COL_CATEGORY]
                    .apply(lambda s: set(s.dropna().unique().tolist()))
                    .reset_index()
                    .rename(columns={COL_CATEGORY: "categories"})
                )

                # Кластеры по объёму и регулярности анализируемого в окне K (как в блоке «Кластерный анализ»)
                available_periods_lc = (max_rank_lc - client_cohort_rank_lc + 1).clip(lower=0, upper=k_int_lc).astype(int)

                per_client_lc = pd.DataFrame({"client_id": sorted(cohort_clients_lc)})
                per_client_lc["cohort_rank"] = per_client_lc["client_id"].map(client_cohort_rank_lc).astype(float)
                per_client_lc["available_periods"] = per_client_lc["client_id"].map(available_periods_lc).fillna(k_int_lc).astype(int)
                df_analyzable_lc = df_purchases_lc[df_purchases_lc[COL_CATEGORY].isin(analyzable_list)].copy()
                df_analyzable_lc = df_analyzable_lc.merge(per_client_lc[["client_id", "cohort_rank"]], on="client_id", how="inner")
                df_analyzable_lc["_in_window"] = (df_analyzable_lc["period_rank"] >= df_analyzable_lc["cohort_rank"]) & (df_analyzable_lc["period_rank"] < df_analyzable_lc["cohort_rank"] + k_int_lc)
                df_analyzable_lc = df_analyzable_lc[df_analyzable_lc["_in_window"]]
                if not df_analyzable_lc.empty:
                    agg_lc = (
                        df_analyzable_lc.groupby("client_id")
                        .agg(volume=(COL_QUANTITY, "sum"), active_periods=("period_rank", "nunique"))
                        .reset_index()
                    )
                    per_client_lc = per_client_lc.merge(agg_lc, on="client_id", how="left")
                if "volume" not in per_client_lc.columns:
                    per_client_lc["volume"] = 0
                else:
                    per_client_lc["volume"] = per_client_lc["volume"].fillna(0).astype(int)
                if "active_periods" not in per_client_lc.columns:
                    per_client_lc["active_periods"] = 0
                else:
                    per_client_lc["active_periods"] = per_client_lc["active_periods"].fillna(0).astype(int)
                denom_lc = per_client_lc["available_periods"].replace(0, 1)
                per_client_lc["regularity"] = (per_client_lc["active_periods"] / denom_lc).clip(0, 1).astype(float)
                per_client_lc["cluster"] = "Не покупали"
                df_fit_lc = per_client_lc[per_client_lc["volume"] > 0].copy()
                if not df_fit_lc.empty:
                    v33_lc = float(df_fit_lc["volume"].quantile(1 / 3))
                    v67_lc = float(df_fit_lc["volume"].quantile(2 / 3))
                    r33, r67 = 1 / 3, 2 / 3

                    def _assign_cluster_lc(row):
                        v, r = row["volume"], row["regularity"]
                        if v >= v67_lc:
                            if r >= r67:
                                return "Активные (VIP)"
                            if r >= r33:
                                return "Регулярные с высоким объёмом"
                            return "Разовая крупная покупка"
                        if v >= v33_lc:
                            if r >= r67:
                                return "Средняя активность"
                            if r >= r33:
                                return "Средняя активность"
                            return "Крупные нерегулярные"
                        if r >= r67:
                            return "Периодические (малый объём)"
                        if r >= r33:
                            return "Низкая активность"
                        return "Разовая покупка"

                    df_fit_lc["cluster"] = df_fit_lc.apply(_assign_cluster_lc, axis=1)
                    per_client_lc = per_client_lc.merge(df_fit_lc[["client_id", "cluster"]], on="client_id", how="left", suffixes=("", "_fit"))
                    per_client_lc["cluster"] = per_client_lc["cluster_fit"].fillna(per_client_lc["cluster"])
                    per_client_lc = per_client_lc.drop(columns=["cluster_fit"], errors="ignore")

                display_to_full_lc = {_cluster_display_name(c): c for c in CLUSTER_8_ORDER}
                display_to_full_lc["Не покупали"] = "Не покупали"
                if "Все кластеры" in selected_clusters_lifecycle or not selected_clusters_lifecycle:
                    selected_cluster_set = set(CLUSTER_8_ORDER + ["Не покупали"])
                else:
                    selected_cluster_set = set(display_to_full_lc.get(s, s) for s in selected_clusters_lifecycle)
                cohort_clients_filtered = set(per_client_lc[per_client_lc["cluster"].isin(selected_cluster_set)]["client_id"].tolist())

                # Продажи анализируемого на объём якорного — только по выбранным кластерам
                df1_anchor_lc = df1_with_period.copy()
                df1_anchor_lc["_client_norm"] = _norm_client_id(df1_anchor_lc[COL_CLIENT])
                df1_anchor_lc = df1_anchor_lc[df1_anchor_lc["_client_norm"].isin(cohort_clients_filtered)]
                df1_anchor_lc["_in_window"] = df1_anchor_lc.apply(_in_window_lc, axis=1)
                q_anchor_lc = df1_anchor_lc.loc[df1_anchor_lc["_in_window"], COL_QUANTITY].sum()

                selected_in_doc1_lc = [c for c in selected_categories_lifecycle if c in categories_from_doc1_set]
                selected_in_doc2_lc = [c for c in selected_categories_lifecycle if c in set(categories_from_doc2)]
                parts_an_lc = []
                if selected_in_doc1_lc:
                    d1_lc = df1_with_period[df1_with_period[COL_CATEGORY].isin(selected_in_doc1_lc)].copy()
                    d1_lc["_client_norm"] = _norm_client_id(d1_lc[COL_CLIENT])
                    d1_lc = d1_lc[d1_lc["_client_norm"].isin(cohort_clients_filtered)]
                    d1_lc["_in_window"] = d1_lc.apply(_in_window_lc, axis=1)
                    parts_an_lc.append(d1_lc.loc[d1_lc["_in_window"], [COL_CATEGORY, COL_QUANTITY]])
                if selected_in_doc2_lc:
                    d2_lc = df2_with_period[df2_with_period[COL_CATEGORY].isin(selected_in_doc2_lc)].copy()
                    d2_lc["_client_norm"] = _norm_client_id(d2_lc[COL_CLIENT])
                    d2_lc = d2_lc[d2_lc["_client_norm"].isin(cohort_clients_filtered)]
                    d2_lc["_in_window"] = d2_lc.apply(_in_window_lc, axis=1)
                    parts_an_lc.append(d2_lc.loc[d2_lc["_in_window"], [COL_CATEGORY, COL_QUANTITY]])
                if parts_an_lc:
                    df_an_lc = pd.concat(parts_an_lc, ignore_index=True)
                    q_by_cat_lc = df_an_lc.groupby(COL_CATEGORY)[COL_QUANTITY].sum().reindex(selected_categories_lifecycle).fillna(0).astype(int)
                else:
                    q_by_cat_lc = pd.Series(dtype=int)
                q_analyzed_lc = int(q_by_cat_lc.sum()) if len(q_by_cat_lc) else 0

                if not cohort_clients_filtered:
                    sales_section_html = (
                        f'<span class="block-block-title">Продажи анализируемого товара на объём якорного</span>'
                        f'<p class="block-p">В выбранных кластерах нет клиентов — коэффициент не рассчитан.</p>'
                    )
                elif q_anchor_lc and q_anchor_lc > 0:
                    r_ratio_lc = q_analyzed_lc / q_anchor_lc
                    expected_int_lc = int(round(n_anchor_lc * r_ratio_lc))
                    period_range_caption_sales = format_period_range_for_caption(
                        cohorts_to_use_lc, cohort_ranks, rank_to_period, k_periods_lifecycle, is_months
                    )
                    analyzed_names_lc = (
                        selected_categories_lifecycle[0]
                        if len(selected_categories_lifecycle) == 1
                        else ", ".join(selected_categories_lifecycle)
                    )
                    anchor_esc_lc = category_label.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    analyzable_esc_lc = analyzed_names_lc.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    sales_section_html = (
                        f'<span class="block-block-title">Продажи анализируемого товара на объём якорного</span>'
                        f'<p class="block-p">Объём анализируемого товара на единицу якорного товара: <span class="block-num">{r_ratio_lc:.2f}</span>.</p>'
                    )
                else:
                    sales_section_html = (
                        f'<span class="block-block-title">Продажи анализируемого товара на объём якорного</span>'
                        f'<p class="block-p">В выбранных кластерах и периоде нет покупок якорного товара — коэффициент не рассчитан.</p>'
                    )

                # Сетка «клиент × период» и флаги покупок (для таблицы и текста «Цикл жизни»)
                client_weeks = []
                for c in cohort_clients_lc:
                    r0 = client_cohort_rank_lc.get(c)
                    if r0 is None or pd.isna(r0):
                        continue
                    r0 = int(r0)
                    for t in range(k_int_lc):
                        client_weeks.append({"client_id": c, "t": t, "period_rank": r0 + t})
                df_cw = pd.DataFrame(client_weeks)
                df_cw = df_cw.merge(client_period_cats, on=["client_id", "period_rank"], how="left")
                df_cw["categories"] = df_cw["categories"].apply(_to_set)

                df_cw["bought_anchor"] = df_cw["categories"].apply(lambda s: bool(s & anchor_cats))
                for i, cat in enumerate(analyzable_list):
                    df_cw[f"bought_a{i}"] = df_cw["categories"].apply(lambda s, c=cat: c in s)
                df_cw["bought_other"] = df_cw["categories"].apply(lambda s: bool(s & other_cats))
                df_cw["no_purchase"] = df_cw["categories"].apply(lambda s: len(s) == 0)
                if analyzable_list:
                    df_cw["bought_any_analyzable"] = df_cw[[f"bought_a{i}" for i in range(len(analyzable_list))]].any(axis=1)
                else:
                    df_cw["bought_any_analyzable"] = False

                df_cw = df_cw[df_cw["client_id"].isin(cohort_clients_filtered)]

                # Взаимоисключающий тип покупки (8 категорий, сумма по строке = 100%)
                def _purchase_bucket(row):
                    if row["no_purchase"]:
                        return "no_purchase"
                    a, an, o = row["bought_anchor"], row["bought_any_analyzable"], row["bought_other"]
                    if a and an and o:
                        return "all_three"
                    if a and an:
                        return "anchor_analyzable"
                    if a and o:
                        return "anchor_other"
                    if an and o:
                        return "analyzable_other"
                    if a:
                        return "only_anchor"
                    if an:
                        return "only_analyzable"
                    if o:
                        return "only_other"
                    return "no_purchase"

                df_cw["purchase_bucket"] = df_cw.apply(_purchase_bucket, axis=1)

                BUCKET_ORDER = [
                    "only_anchor",
                    "only_analyzable",
                    "anchor_analyzable",
                    "only_other",
                    "anchor_other",
                    "analyzable_other",
                    "all_three",
                    "no_purchase",
                ]
                BUCKET_LABELS = {
                    "only_anchor": "Только якорный",
                    "only_analyzable": "Только анализируемый",
                    "anchor_analyzable": "Якорный и анализируемый",
                    "only_other": "Только прочие",
                    "anchor_other": "Якорный и прочие",
                    "analyzable_other": "Анализируемый и прочие",
                    "all_three": "Якорный + анализируемый + прочие",
                    "no_purchase": "Нет покупок",
                }

                N_lc = len(cohort_clients_filtered)
                if N_lc == 0:
                    st.warning("В выбранных кластерах нет клиентов. Выберите другие кластеры или когорты.")
                else:
                    agg_d = {"bought_anchor": ("bought_anchor", "sum"), "bought_other": ("bought_other", "sum"), "no_purchase": ("no_purchase", "sum"), "bought_any_analyzable": ("bought_any_analyzable", "sum")}
                    for i in range(len(analyzable_list)):
                        agg_d[f"bought_a{i}"] = (f"bought_a{i}", "sum")
                    summary_by_week = df_cw.groupby("t").agg(**agg_d).reset_index()

                    # Сводка по (t, purchase_bucket) для HTML-таблицы
                    bucket_counts = df_cw.groupby(["t", "purchase_bucket"]).size().unstack(fill_value=0)
                    for b in BUCKET_ORDER:
                        if b not in bucket_counts.columns:
                            bucket_counts[b] = 0
                    bucket_counts = bucket_counts.reindex(columns=BUCKET_ORDER).fillna(0).astype(int)

                    half_life_50_t = None
                    half_life_30_t = None
                    half_life_10_t = None
                    pct_before_50 = None
                    pct_at_50 = None
                    pct_at_30 = None
                    pct_at_10 = None
                    for _, r in summary_by_week.iterrows():
                        pct = 100 * r["bought_any_analyzable"] / N_lc if N_lc else 0
                        t = int(r["t"])
                        if half_life_50_t is None and pct < 50:
                            half_life_50_t = t
                            pct_at_50 = pct
                            row_before = summary_by_week[summary_by_week["t"] == t - 1]
                            pct_before_50 = 100 * row_before.iloc[0]["bought_any_analyzable"] / N_lc if t > 0 and not row_before.empty else None
                        if half_life_30_t is None and pct < 30:
                            half_life_30_t = t
                            pct_at_30 = pct
                        if half_life_10_t is None and pct < 10:
                            half_life_10_t = t
                            pct_at_10 = pct
                        if half_life_50_t is not None and half_life_30_t is not None and half_life_10_t is not None:
                            break
                    half_life_per_product = []
                    for i in range(len(analyzable_list)):
                        hi_50, hi_30, hi_10 = None, None, None
                        pa_50, pa_30, pa_10 = None, None, None
                        pb_50 = None
                        for _, r in summary_by_week.iterrows():
                            pct = 100 * r[f"bought_a{i}"] / N_lc if N_lc else 0
                            t = int(r["t"])
                            if hi_50 is None and pct < 50:
                                hi_50 = t
                                pa_50 = pct
                                row_b = summary_by_week[summary_by_week["t"] == t - 1]
                                pb_50 = 100 * row_b.iloc[0][f"bought_a{i}"] / N_lc if t > 0 and not row_b.empty else None
                            if hi_30 is None and pct < 30:
                                hi_30 = t
                                pa_30 = pct
                            if hi_10 is None and pct < 10:
                                hi_10 = t
                                pa_10 = pct
                        half_life_per_product.append({
                            "name": analyzable_list[i],
                            "w50": (hi_50 + 1) if hi_50 is not None else None,
                            "pct_bef50": pb_50,
                            "pct_at_50": pa_50,
                            "w30": (hi_30 + 1) if hi_30 is not None else None,
                            "pct_at_30": pa_30,
                            "w10": (hi_10 + 1) if hi_10 is not None else None,
                            "pct_at_10": pa_10,
                        })
                    df_cw_sorted = df_cw.sort_values(["client_id", "t"])
                    first_miss = df_cw_sorted[df_cw_sorted["bought_any_analyzable"] == False].groupby("client_id")["t"].min().reset_index().rename(columns={"t": "first_miss"})
                    clients_all = df_cw_sorted["client_id"].unique()
                    consec = pd.DataFrame({"client_id": clients_all}).merge(first_miss, on="client_id", how="left")
                    consec["consecutive_weeks"] = consec["first_miss"].fillna(k_int_lc).astype(int)
                    avg_consecutive_weeks = consec["consecutive_weeks"].mean() if len(consec) else 0.0
                    median_consecutive_weeks = consec["consecutive_weeks"].median() if len(consec) else 0.0

                    gap_lengths = []
                    for cid in df_cw["client_id"].unique():
                        seq = df_cw[df_cw["client_id"] == cid].set_index("t").reindex(range(k_int_lc)).fillna(False)["bought_any_analyzable"].tolist()
                        i = 0
                        while i < k_int_lc:
                            if not seq[i]:
                                j = i
                                while j < k_int_lc and not seq[j]:
                                    j += 1
                                gap_lengths.append(j - i)
                                i = j
                            else:
                                i += 1
                    median_gap = float(np.median(gap_lengths)) if gap_lengths else 1.0
                    sustained_threshold = max(1, int(round(median_gap)))
                    first_sustained_start = {}
                    first_sustained_end = {}
                    first_sustained_other = {}
                    first_sustained_none = {}
                    for cid in df_cw["client_id"].unique():
                        rows = df_cw[df_cw["client_id"] == cid].sort_values("t")
                        seq = rows.set_index("t").reindex(range(k_int_lc)).fillna(False)["bought_any_analyzable"].tolist()
                        i = 0
                        found = None
                        while i < k_int_lc:
                            if not seq[i]:
                                j = i
                                while j < k_int_lc and not seq[j]:
                                    j += 1
                                if (j - i) >= sustained_threshold:
                                    found = (i, j - i)
                                    break
                                i = j
                            else:
                                i += 1
                        if found is not None:
                            t_start, gap_len = found
                            first_sustained_start[cid] = t_start
                            first_sustained_end[cid] = t_start + gap_len
                            window = df_cw[(df_cw["client_id"] == cid) & (df_cw["t"] >= t_start) & (df_cw["t"] < t_start + gap_len)]
                            first_sustained_other[cid] = window["bought_other"].any()
                            first_sustained_none[cid] = window["no_purchase"].any()
                    n_sustained = len(first_sustained_start)
                    avg_first_sustained_week = np.mean(list(first_sustained_start.values())) if first_sustained_start else None
                    pct_in_gap_other = 100 * sum(first_sustained_other.values()) / n_sustained if n_sustained else 0
                    pct_in_gap_none = 100 * sum(first_sustained_none.values()) / n_sustained if n_sustained else 0
                    pct_clients_with_sustained = 100 * n_sustained / N_lc if N_lc else 0
                    other_cat_count_in_gap = {}
                    for cid in first_sustained_start:
                        t_start, t_end = first_sustained_start[cid], first_sustained_end.get(cid, first_sustained_start[cid] + 1)
                        window = df_cw[(df_cw["client_id"] == cid) & (df_cw["t"] >= t_start) & (df_cw["t"] < t_end)]
                        for _, r in window.iterrows():
                            for c in (r["categories"] & other_cats):
                                other_cat_count_in_gap.setdefault(c, set()).add(cid)
                    top3_other_in_gap = sorted(
                        [(c, len(s)) for c, s in other_cat_count_in_gap.items()],
                        key=lambda x: -x[1]
                    )[:3]
                    top3_other_in_gap_pct = [(c, 100 * cnt / N_lc) for c, cnt in top3_other_in_gap] if N_lc else []

                    last_purchase_week = df_cw[df_cw["bought_any_analyzable"]].groupby("client_id")["t"].max()
                    last_pw = last_purchase_week.reindex(consec["client_id"].values)
                    last_pw.index = consec.index
                    exited_mask = (last_pw < k_int_lc - 1) | last_pw.isna()
                    exited_clients = consec.loc[exited_mask]["client_id"].tolist()
                    pct_exited = 100 * len(exited_clients) / N_lc if N_lc else 0
                    avg_last_purchase_week = last_pw.loc[exited_mask].dropna().mean() if exited_mask.any() else None
                    exit_per_product = []
                    for i in range(len(analyzable_list)):
                        last_pi = df_cw[df_cw[f"bought_a{i}"]].groupby("client_id")["t"].max()
                        last_pi = last_pi.reindex(consec["client_id"].values)
                        last_pi.index = consec.index
                        exited_i = (last_pi < k_int_lc - 1) | last_pi.isna()
                        pct_exited_i = 100 * exited_i.sum() / N_lc if N_lc else 0.0
                        avg_last_i = last_pi.loc[exited_i].dropna().mean() if exited_i.any() else None
                        exit_per_product.append({"pct": pct_exited_i, "avg_week": avg_last_i})
                    last_anchor_week = df_cw[df_cw["bought_anchor"]].groupby("client_id")["t"].max()
                    last_anchor_pw = last_anchor_week.reindex(consec["client_id"].values)
                    last_anchor_pw.index = consec.index
                    exited_anchor_mask = (last_anchor_pw < k_int_lc - 1) | last_anchor_pw.isna()
                    pct_exited_anchor = 100 * exited_anchor_mask.sum() / N_lc if N_lc else 0.0
                    avg_last_anchor_week = last_anchor_pw.loc[exited_anchor_mask].dropna().mean() if exited_anchor_mask.any() else None

                    t_mid = (k_int_lc - 1) // 2
                    mid_rows = summary_by_week[summary_by_week["t"] == t_mid]
                    row_mid = mid_rows.iloc[0] if len(mid_rows) else None
                    row_0 = summary_by_week[summary_by_week["t"] == 0].iloc[0] if len(summary_by_week[summary_by_week["t"] == 0]) else None
                    pct_anchor_mid = 100 * row_mid["bought_anchor"] / N_lc if row_mid is not None else (100 * row_0["bought_anchor"] / N_lc if row_0 is not None else 0)
                    pct_analyzable_mid = 100 * row_mid["bought_any_analyzable"] / N_lc if row_mid is not None and analyzable_list else None
                    pct_analyzable_first = 100 * row_0["bought_any_analyzable"] / N_lc if analyzable_list and row_0 is not None else None
                    if analyzable_list and N_lc and k_int_lc >= 3:
                        df_weeks_2_3 = df_cw[(df_cw["t"].isin([1, 2])) & (df_cw["bought_any_analyzable"])]
                        pct_analyzable_weeks_2_3 = 100 * df_weeks_2_3["client_id"].nunique() / N_lc
                    else:
                        pct_analyzable_weeks_2_3 = None

                    n_last_weeks = min(3, k_int_lc)
                    t_end_from = k_int_lc - n_last_weeks + 1
                    t_end_to = k_int_lc

                    df_last_week = df_cw[df_cw["t"] == k_int_lc - 1]
                    df_last_n = df_cw[df_cw["t"] >= k_int_lc - n_last_weeks]
                    other_cat_count = {}
                    for _, r in df_last_week.iterrows():
                        for c in (r["categories"] & other_cats):
                            other_cat_count[c] = other_cat_count.get(c, 0) + 1
                    most_popular_other = max(other_cat_count, key=other_cat_count.get) if other_cat_count else None
                    pct_most_popular_other = 100 * other_cat_count.get(most_popular_other, 0) / N_lc if most_popular_other else 0.0

                    client_other_cats_n = {}
                    for _, r in df_last_n.iterrows():
                        if r["bought_other"] and r["categories"] & other_cats:
                            for c in (r["categories"] & other_cats):
                                client_other_cats_n.setdefault(r["client_id"], set()).add(c)
                    other_clients_count_n = {}
                    for cid, cats in client_other_cats_n.items():
                        for c in cats:
                            other_clients_count_n[c] = other_clients_count_n.get(c, set()) | {cid}
                    top3_other = sorted(
                        [(c, len(s)) for c, s in other_clients_count_n.items()],
                        key=lambda x: -x[1]
                    )[:3]
                    top3_other_pct = [(c, 100 * cnt / N_lc) for c, cnt in top3_other] if N_lc else []

                    no_purchase_per_client_n = df_last_n.groupby("client_id")["no_purchase"].all()
                    clients_all_weeks_in_window = df_last_n.groupby("client_id").size() == n_last_weeks
                    clients_none_last_n = (no_purchase_per_client_n & clients_all_weeks_in_window).sum()
                    pct_none_last_n = 100 * clients_none_last_n / N_lc if N_lc else 0.0
                    clients_other_last_n = df_last_n[df_last_n["bought_other"]]["client_id"].nunique()
                    pct_other_last_n = 100 * clients_other_last_n / N_lc if N_lc else 0.0

                    period_unit_lc = "месяц" if is_months else "неделя"
                    period_unit_plural = "месяцев" if is_months else "недель"
                    period_unit_single = "неделя" if not is_months else "месяц"
                    if is_months:
                        n_last_word = "месяц" if n_last_weeks == 1 else ("месяца" if n_last_weeks <= 4 else "месяцев")
                    else:
                        n_last_word = "неделю" if n_last_weeks == 1 else ("недели" if n_last_weeks <= 4 else "недель")
                    end_period_weeks_str = f"{period_unit_single} {t_end_from}–{t_end_to}" if n_last_weeks > 1 else f"{period_unit_single} {t_end_to}"

                    # HTML-таблица: типы покупки (сумма по строке = 100%), первый столбец фиксирован, подсветка преобладающего типа
                    period_unit_single_lc = "неделя" if not is_months else "месяц"
                    bucket_counts_reindexed = bucket_counts.reindex(range(k_int_lc)).fillna(0).astype(int)

                    def _esc(s):
                        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

                    predominant_col_by_t = {}
                    for t in range(k_int_lc):
                        if t not in bucket_counts_reindexed.index:
                            predominant_col_by_t[t] = -1
                            continue
                        row_vals = bucket_counts_reindexed.loc[t, BUCKET_ORDER].values
                        max_idx = int(np.argmax(row_vals))
                        predominant_col_by_t[t] = max_idx

                    thead_cells = [
                        f'<th class="lc-th lc-th-period">{_esc("Период от когорты⬇/Тип покупки⮕")}</th>'
                    ]
                    for b in BUCKET_ORDER:
                        thead_cells.append(f'<th class="lc-th">{_esc(BUCKET_LABELS[b])}</th>')

                    tbody_rows = []
                    pastel_row_classes = ["lc-row-0", "lc-row-1", "lc-row-2", "lc-row-3", "lc-row-4", "lc-row-5"]
                    for t in range(k_int_lc):
                        pred_idx = predominant_col_by_t.get(t, -1)
                        period_label = f"{period_unit_single_lc} {t + 1}"
                        row_class = pastel_row_classes[t % len(pastel_row_classes)]
                        cells_html = [f'<td class="lc-td lc-td-period {row_class}">{_esc(period_label)}</td>']
                        for col_idx, b in enumerate(BUCKET_ORDER):
                            cnt = int(bucket_counts_reindexed.loc[t, b]) if t in bucket_counts_reindexed.index else 0
                            pct = 100 * cnt / N_lc if N_lc else 0
                            pred_class = " lc-predominant" if col_idx == pred_idx else ""
                            cells_html.append(f'<td class="lc-td {row_class}{pred_class}">{cnt} ({pct:.1f}%)</td>')
                        tbody_rows.append(f'<tr class="{row_class}">' + "".join(cells_html) + "</tr>")

                    # Пояснение к таблице — над таблицей, построчно
                    about_table_html = (
                        '<div class="block-about-table">'
                        "<strong>О таблице :</strong>"
                        "<div class=\"block-about-line\">Строки — периоды с начала когорты.</div>"
                        "<div class=\"block-about-line\">Столбцы — тип покупки в этот период. Каждый клиент попадает ровно в один тип, поэтому по строке сумма = 100%.</div>"
                        "<div class=\"block-about-line\">Зелёная ячейка — преобладающий тип в этот период.</div>"
                        "</div>"
                    )
                    about_table_css = (
                        '<style>'
                        '.block-about-table { background: #1a1a1a; border: 1px solid #333; color: #e8e8e8; '
                        'padding: 0.6rem 0.75rem; margin: 0.5rem 0 0.35rem 0; font-size: 0.9rem; line-height: 1.5; border-radius: 6px; }'
                        '.block-about-table .block-about-line { margin-bottom: 0.25rem; }'
                        '.block-about-table .block-about-line:last-child { margin-bottom: 0; }'
                        '</style>'
                    )
                    st.markdown(about_table_css + about_table_html, unsafe_allow_html=True)

                    lc_table_html = (
                        '<div class="lc-table-wrapper">'
                        '<table class="lc-table">'
                        "<thead><tr>" + "".join(thead_cells) + "</tr></thead>"
                        "<tbody>" + "".join(tbody_rows) + "</tbody>"
                        "</table></div>"
                    )

                    lc_table_css = """
                    <style>
                    .lc-table-wrapper { overflow-x: auto; overflow-y: visible; margin: 0.5rem 0; max-width: 100%%; }
                    .lc-table { border-collapse: collapse; width: 100%%; font-size: 0.8rem; background: #0d0d0d; color: #e8e8e8; table-layout: fixed; }
                    .lc-table thead { position: sticky; top: 0; z-index: 4; }
                    .lc-table thead tr { background: #1a1a1a; }
                    .lc-th, .lc-td { border: 1px solid #333; padding: 0.35rem 0.45rem; text-align: right; }
                    .lc-th-period, .lc-td-period { position: sticky; left: 0; z-index: 2; text-align: left; font-weight: 600; }
                    .lc-th-period { z-index: 5; background: #1a1a1a; }
                    .lc-table tbody .lc-td-period { background: inherit; }
                    .lc-table thead .lc-th-period { box-shadow: 2px 0 4px rgba(0,0,0,0.3); }
                    .lc-table tbody .lc-td-period { box-shadow: 2px 0 4px rgba(0,0,0,0.2); }
                    .lc-row-0 .lc-td:not(.lc-td-period) { background: rgba(255, 218, 224, 0.12); }
                    .lc-row-0 .lc-td-period { background: rgba(255, 218, 224, 0.18); }
                    .lc-row-1 .lc-td:not(.lc-td-period) { background: rgba(218, 230, 255, 0.12); }
                    .lc-row-1 .lc-td-period { background: rgba(218, 230, 255, 0.18); }
                    .lc-row-2 .lc-td:not(.lc-td-period) { background: rgba(218, 255, 230, 0.12); }
                    .lc-row-2 .lc-td-period { background: rgba(218, 255, 230, 0.18); }
                    .lc-row-3 .lc-td:not(.lc-td-period) { background: rgba(255, 236, 218, 0.12); }
                    .lc-row-3 .lc-td-period { background: rgba(255, 236, 218, 0.18); }
                    .lc-row-4 .lc-td:not(.lc-td-period) { background: rgba(238, 218, 255, 0.12); }
                    .lc-row-4 .lc-td-period { background: rgba(238, 218, 255, 0.18); }
                    .lc-row-5 .lc-td:not(.lc-td-period) { background: rgba(218, 255, 248, 0.12); }
                    .lc-row-5 .lc-td-period { background: rgba(218, 255, 248, 0.18); }
                    .lc-predominant { background: rgba(200, 255, 220, 0.35) !important; font-weight: 600; }
                    </style>
                    """
                    st.markdown(lc_table_css + lc_table_html, unsafe_allow_html=True)

                    last = summary_by_week.iloc[-1]
                    pct_anchor_last = 100 * last["bought_anchor"] / N_lc
                    pct_other_last = 100 * last["bought_other"] / N_lc
                    pct_none_last = 100 * last["no_purchase"] / N_lc
                    pct_analyzable_last = [100 * last[f"bought_a{i}"] / N_lc for i in range(len(analyzable_list))]
                    first_row = summary_by_week.iloc[0]
                    pct_anchor_first = 100 * first_row["bought_anchor"] / N_lc

                    period_range_caption_lc = format_period_range_for_caption(
                        cohorts_to_use_lc, cohort_ranks, rank_to_period, k_periods_lifecycle, is_months
                    )
                    period_word_until = "недели" if not is_months else "месяца"
                    period_word_on = "неделе" if not is_months else "месяце"
                    analyzable_names_esc = ", ".join([c.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;") for c in analyzable_list])
                    def _one_half_life_lines(names_esc, w50, pct_bef50, pct_at_50, w30, pct_at_30, w10, pct_at_10, fallback_text):
                        lines = []
                        if w50 is not None:
                            pct_bef = f"{pct_bef50:.1f}%" if pct_bef50 is not None else "—"
                            if w50 > 1:
                                lines.append(
                                    f"Доля покупающих анализируемые продукты (<span class=\"block-product\">{names_esc}</span>) падает ниже 50% начиная с {period_word_until} <span class=\"block-num\">{w50}</span> "
                                    f"(на {period_word_on} <span class=\"block-num\">{w50 - 1}</span> — <span class=\"block-num\">{pct_bef}</span>, на {period_word_on} <span class=\"block-num\">{w50}</span> — <span class=\"block-num\">{pct_at_50:.1f}%</span>)."
                                )
                            else:
                                lines.append(
                                    f"Доля покупающих анализируемые продукты (<span class=\"block-product\">{names_esc}</span>) падает ниже 50% начиная с {period_word_until} <span class=\"block-num\">1</span> (на {period_word_on} <span class=\"block-num\">1</span> — <span class=\"block-num\">{pct_at_50:.1f}%</span>)."
                                )
                        if w30 is not None and pct_at_30 is not None:
                            lines.append(f"Ниже 30% — с {period_word_until} <span class=\"block-num\">{w30}</span> (<span class=\"block-num\">{pct_at_30:.1f}%</span>).")
                        if w10 is not None and pct_at_10 is not None:
                            lines.append(f"Ниже 10% — с {period_word_until} <span class=\"block-num\">{w10}</span> (<span class=\"block-num\">{pct_at_10:.1f}%</span>).")
                        if not lines:
                            lines.append(fallback_text)
                        return lines
                    half_life_divs = []
                    fallback_overall = f"На всём периоде (<span class=\"block-num\">{k_int_lc}</span> {period_unit_plural}) более половины когорты покупают хотя бы один из анализируемых продуктов (<span class=\"block-product\">{analyzable_names_esc}</span>)."
                    w50 = (half_life_50_t + 1) if half_life_50_t is not None else None
                    w30 = (half_life_30_t + 1) if half_life_30_t is not None else None
                    w10 = (half_life_10_t + 1) if half_life_10_t is not None else None
                    for line in _one_half_life_lines(
                        analyzable_names_esc, w50, pct_before_50, pct_at_50, w30, pct_at_30, w10, pct_at_10, fallback_overall
                    ):
                        half_life_divs.append(f'<div class="block-p4-line">{line}</div>')
                    if len(analyzable_list) > 1:
                        half_life_spacer = '<div class="block-p4-line block-spacer"></div>'
                        for hp in half_life_per_product:
                            half_life_divs.append(half_life_spacer)
                            cat_esc = hp["name"].replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                            fallback_cat = f"На всём периоде более половины когорты покупают <span class=\"block-product\">{cat_esc}</span>."
                            for line in _one_half_life_lines(
                                cat_esc, hp["w50"], hp["pct_bef50"], hp["pct_at_50"],
                                hp["w30"], hp["pct_at_30"], hp["w10"], hp["pct_at_10"], fallback_cat
                            ):
                                half_life_divs.append(f'<div class="block-p4-line">{line}</div>')
                    half_life_text = "".join(half_life_divs)

                    anchor_name_esc = category_label.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    week_1 = 1
                    week_mid = t_mid + 1
                    week_end = k_int_lc
                    cohort_short_names = [lb.split(" (")[0] for lb in cohorts_to_use_lc]
                    if len(cohort_short_names) == 1:
                        cohorts_list_str = cohort_short_names[0]
                    else:
                        cohorts_list_str = f"{cohort_short_names[0]}-{cohort_short_names[-1]}"
                    n_c = len(cohorts_to_use_lc)
                    if n_c % 10 == 1 and n_c != 11:
                        cohort_word = "когорта"
                    elif n_c % 10 in (2, 3, 4) and n_c not in (12, 13, 14):
                        cohort_word = "когорты"
                    else:
                        cohort_word = "когорт"
                    header_first_line = (
                        f"{period_range_caption_lc}; <span class=\"block-num\">{n_c}</span> {cohort_word} (<span class=\"block-product\">{cohorts_list_str}</span>); "
                        f"<span class=\"block-num\">{N_lc}</span> клиентов; Первые <span class=\"block-num\">{k_int_lc}</span> {period_unit_plural} с момента когорты."
                    )
                    p1_anchor_body = (
                        f'<span class="block-product">{anchor_name_esc}</span>: '
                        f'<ul class="block-ul">'
                        f'<li>{period_unit_single} <span class="block-num">{week_1}</span> — <span class="block-num">{pct_anchor_first:.1f}%</span></li>'
                        f'<li>Середина ({period_unit_single} <span class="block-num">{week_mid}</span>) — <span class="block-num">{pct_anchor_mid:.1f}%</span></li>'
                        f'<li>К концу ({period_unit_single} <span class="block-num">{week_end}</span>) — <span class="block-num">{pct_anchor_last:.1f}%</span></li>'
                        f'</ul>'
                    )
                    p2_analyzable_lines = []
                    if analyzable_list:
                        pct_analyzable_end_overall = 100 * last["bought_any_analyzable"] / N_lc
                        first_period_phrase = "На первой неделе" if not is_months else "На 1 месяца"
                        period_loc_single = "неделе" if not is_months else "месяце"
                        products_label = ", ".join([c.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;") for c in analyzable_list])
                        p2_analyzable_lines.append(
                            f'<div class="block-p4-line">Доля от когорты присутствует в анализируемых товарах (<span class="block-product">{products_label}</span>).</div>'
                        )
                        p2_analyzable_lines.append(
                            f'<div class="block-p4-line">{first_period_phrase} — <span class="block-num">{pct_analyzable_first:.1f}%</span> клиентов когорты</div>'
                        )
                        if pct_analyzable_weeks_2_3 is not None:
                            p2_analyzable_lines.append(f'<div class="block-p4-line">Недели 2–3: <span class="block-num">{pct_analyzable_weeks_2_3:.1f}%</span></div>')
                        p2_analyzable_lines.append(
                            f'<div class="block-p4-line">В середине периода на <span class="block-num">{week_mid}</span> {period_loc_single} — <span class="block-num">{pct_analyzable_mid:.1f}%</span></div>'
                        )
                        p2_analyzable_lines.append(
                            f'<div class="block-p4-line">К концу на <span class="block-num">{week_end}</span> {period_loc_single} — <span class="block-num">{pct_analyzable_end_overall:.1f}%</span> клиентов когорты.</div>'
                        )
                        if len(analyzable_list) > 1:
                            spacer_line = '<div class="block-p4-line block-spacer"></div>'
                            for i, cat in enumerate(analyzable_list):
                                p2_analyzable_lines.append(spacer_line)
                                cat_esc = cat.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                                rows_0 = summary_by_week[summary_by_week["t"] == 0]
                                rows_mid = summary_by_week[summary_by_week["t"] == t_mid]
                                row_0_cat = rows_0.iloc[0] if len(rows_0) else None
                                row_mid_cat = rows_mid.iloc[0] if len(rows_mid) else None
                                pct_first_cat = 100 * row_0_cat[f"bought_a{i}"] / N_lc if row_0_cat is not None else 0.0
                                pct_mid_cat = 100 * row_mid_cat[f"bought_a{i}"] / N_lc if row_mid_cat is not None else 0.0
                                pct_end_cat = 100 * last[f"bought_a{i}"] / N_lc
                                df_23_cat = df_cw[(df_cw["t"].isin([1, 2])) & (df_cw[f"bought_a{i}"])] if k_int_lc >= 3 else pd.DataFrame()
                                pct_23_cat = 100 * df_23_cat["client_id"].nunique() / N_lc if not df_23_cat.empty and N_lc else None
                                p2_analyzable_lines.append(
                                    f'<div class="block-p4-line">Доля от когорты присутствует в анализируемых товарах (<span class="block-product">{cat_esc}</span>).</div>'
                                )
                                p2_analyzable_lines.append(
                                    f'<div class="block-p4-line">{first_period_phrase} — <span class="block-num">{pct_first_cat:.1f}%</span> клиентов когорты</div>'
                                )
                                if pct_23_cat is not None:
                                    p2_analyzable_lines.append(f'<div class="block-p4-line">Недели 2–3: <span class="block-num">{pct_23_cat:.1f}%</span></div>')
                                p2_analyzable_lines.append(
                                    f'<div class="block-p4-line">В середине периода на <span class="block-num">{week_mid}</span> {period_loc_single} — <span class="block-num">{pct_mid_cat:.1f}%</span></div>'
                                )
                                p2_analyzable_lines.append(
                                    f'<div class="block-p4-line">К концу на <span class="block-num">{week_end}</span> {period_loc_single} — <span class="block-num">{pct_end_cat:.1f}%</span> клиентов когорты.</div>'
                                )
                    p2_analyzable_html = "".join(p2_analyzable_lines) if p2_analyzable_lines else ""
                    p2_outcomes_html = (
                        f"За последние <span class=\"block-num\">{n_last_weeks}</span> {n_last_word} ({end_period_weeks_str}): "
                        f'<ul class="block-ul">'
                        f'<li><span class="block-num">{pct_other_last_n:.1f}%</span> — покупали прочие категории (без анализируемого)</li>'
                        f'<li><span class="block-num">{pct_none_last_n:.1f}%</span> — не имели покупок ни в одну из этих {n_last_word}</li>'
                        f'</ul>'
                    )
                    p2_other_popular_html = ""
                    if top3_other_pct:
                        top3_items = "".join([
                            f'<li><span class="block-product">{c.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")}</span> — <span class="block-num">{pct:.1f}%</span> когорты</li>'
                            for c, pct in top3_other_pct
                        ])
                        p2_other_popular_html = (
                            f"Топ-3 прочих категорий за последние <span class=\"block-num\">{n_last_weeks}</span> {n_last_word}: "
                            f'<ul class="block-ul">{top3_items}</ul>'
                        )
                    p3_html = half_life_text
                    period_loc = "неделе" if not is_months else "месяце"
                    period_loc_gen = "недели" if not is_months else "месяца"
                    period_one = "одна неделя" if not is_months else "один месяц"
                    p4_lines = []
                    if analyzable_list:
                        p4_lines.append(
                            f'<div class="block-p4-line"><strong>Типичный перерыв:</strong> медиана между покупками анализируемого продукта — <span class="block-num">{median_gap:.1f}</span> {period_unit_plural}.</div>'
                        )
                        if n_sustained > 0 and avg_first_sustained_week is not None:
                            avg_first_sustained_week_1based = avg_first_sustained_week + 1
                            p4_lines.append(
                                f'<div class="block-p4-line">У <span class="block-num">{pct_clients_with_sustained:.1f}%</span> клиентов когорты первый перерыв больше, чем <span class="block-num">{sustained_threshold}</span> {period_loc_gen}, в среднем с {period_loc_gen} <span class="block-num">{avg_first_sustained_week_1based:.1f}</span>.</div>'
                            )
                            gap_other_top3 = "".join([
                                f'<li><span class="block-product">{c.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")}</span> — <span class="block-num">{pct:.1f}%</span> когорты</li>'
                                for c, pct in top3_other_in_gap_pct
                            ]) if top3_other_in_gap_pct else ""
                            p4_lines.append(
                                f'<div class="block-p4-line">В перерыве: <span class="block-num">{pct_in_gap_none:.1f}%</span> — без покупок вообще; <span class="block-num">{pct_in_gap_other:.1f}%</span> — покупали прочие.</div>'
                            )
                            if gap_other_top3:
                                p4_lines.append(f'<div class="block-p4-line">Топ-3 прочих в перерыве: <ul class="block-ul">{gap_other_top3}</ul></div>')
                        exit_anchor_line = f"Полный уход из якорного продукта: <span class=\"block-num\">{pct_exited_anchor:.1f}%</span> когорты"
                        if avg_last_anchor_week is not None and not np.isnan(avg_last_anchor_week):
                            avg_last_anchor_1based = avg_last_anchor_week + 1
                            exit_anchor_line += f"; в среднем последняя покупка — на {period_loc} <span class=\"block-num\">{avg_last_anchor_1based:.1f}</span>."
                        else:
                            exit_anchor_line += "."
                        p4_lines.append(f'<div class="block-p4-line"><strong>Уход:</strong></div>')
                        p4_lines.append(f'<div class="block-p4-line">{exit_anchor_line}</div>')
                        exit_line_overall = f"Полный уход из анализируемого продукта: <span class=\"block-num\">{pct_exited:.1f}%</span> когорты"
                        if avg_last_purchase_week is not None and not np.isnan(avg_last_purchase_week):
                            avg_last_week_1based = avg_last_purchase_week + 1
                            exit_line_overall += f"; в среднем последняя покупка — на {period_loc} <span class=\"block-num\">{avg_last_week_1based:.1f}</span>."
                        else:
                            exit_line_overall += "."
                        p4_lines.append(f'<div class="block-p4-line">{exit_line_overall}</div>')
                        if len(analyzable_list) > 1:
                            for i, ep in enumerate(exit_per_product):
                                cat_esc = analyzable_list[i].replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                                pct_i = ep["pct"]
                                avg_week_i = ep["avg_week"]
                                exit_line_i = f"Полный уход из анализируемого продукта (<span class=\"block-product\">{cat_esc}</span>): <span class=\"block-num\">{pct_i:.1f}%</span> когорты"
                                if avg_week_i is not None and not np.isnan(avg_week_i):
                                    exit_line_i += f"; в среднем последняя покупка — на {period_loc} <span class=\"block-num\">{avg_week_i + 1:.1f}</span>."
                                else:
                                    exit_line_i += "."
                                p4_lines.append(f'<div class="block-p4-line">{exit_line_i}</div>')
                    p4_html = "".join(p4_lines) if p4_lines else ""

                    lifecycle_box_css = (
                        "<style>"
                        ".block-result-box { background: #0d0d0d; border: 1px solid #333; border-radius: 8px; padding: 1rem 1.25rem; margin: 0.5rem 0; color: #e8e8e8; }"
                        ".block-result-box .block-divider { border-top: 1px solid #333; margin: 0.75rem 0; }"
                        ".block-result-box .block-block-title { font-size: 1.05rem; font-weight: 700; color: #e8e8e8; display: block; margin-bottom: 0.5rem; padding: 0.5rem 0.6rem; border: 1px solid #333; border-radius: 6px; margin-top: 0; background: #1a1a1a; }"
                        ".block-result-box .block-block-title:first-of-type { margin-top: 0; }"
                        ".block-result-box .block-sales-block { margin-bottom: 0.5rem; }"
                        ".block-result-box .block-lifecycle-block { padding-top: 1rem; margin-top: 0.5rem; }"
                        ".block-result-box .block-section-title { font-style: italic; font-weight: 600; margin-top: 1.5rem; margin-bottom: 0.35rem; color: #c4b5fd; display: block; font-size: 1.1rem; }"
                        ".block-result-box .block-section-title:first-of-type { margin-top: 0; }"
                        ".block-period-above { text-align: center; font-weight: 700; font-size: 1.05rem; letter-spacing: 0.02em; color: #e8e8e8; margin-bottom: 0.5rem; }"
                        ".block-period-above .block-num { color: #f0a050; font-weight: bold; }"
                        ".block-period-above .block-product { font-style: italic; color: #e0e0e0; }"
                        ".block-result-box .block-p4-line { margin-bottom: 0.5rem; line-height: 1.45; }"
                        ".block-result-box .block-p4-line:last-child { margin-bottom: 0; }"
                        ".block-result-box .block-p4-line.block-spacer { height: 1.45em; margin-bottom: 0.5rem; min-height: 1px; }"
                        ".block-result-box .block-num { color: #f0a050; font-weight: bold; }"
                        ".block-result-box .block-product { font-style: italic; background: rgba(255,255,255,0.08); color: #e0e0e0; padding: 0.1em 0.35em; border-radius: 4px; border: 1px solid #444; }"
                        ".block-result-box p.block-p, .block-result-box div.block-p { margin: 0 0 0.5rem 0; font-size: 0.95rem; line-height: 1.45; color: #e8e8e8; }"
                        ".block-result-box .block-ul { margin: 0.25rem 0 0.5rem 1.25rem; padding-left: 0.5rem; color: #e8e8e8; }"
                        ".block-result-box .block-ul li { margin-bottom: 0.2rem; }"
                        "</style>"
                    )
                    # Подпись периода — над ячейкой вывода; затем ячейка без подписи внутри
                    lifecycle_box_html = (
                        lifecycle_box_css
                        + f'<div class="block-period-above">{header_first_line}</div>'
                        + f'<div class="block-result-box">'
                        + f'<div class="block-sales-block">'
                        + sales_section_html
                    )
                    if q_anchor_lc and q_anchor_lc > 0:
                        lifecycle_box_html += (
                            f'<p class="block-p">При продаже <span class="block-num">{n_anchor_lc}</span> ед. <span class="block-product">{anchor_esc_lc}</span> в течении '
                            f'<span class="block-num">{k_periods_lifecycle}</span> {period_word} будет продано '
                            f'<span class="block-num">{expected_int_lc}</span> ед. <span class="block-product">{analyzable_esc_lc}</span>.</p>'
                        )
                    lifecycle_box_html += "</div>"
                    lifecycle_box_html += (
                        f'<div class="block-divider"></div>'
                        + f'<div class="block-lifecycle-block">'
                        + f'<span class="block-block-title">Цикл жизни клиента</span>'
                        + f'<span class="block-section-title">Якорный продукт</span>'
                        + f'<p class="block-p">{p1_anchor_body}</p>'
                    )
                    if p2_analyzable_html:
                        lifecycle_box_html += f'<span class="block-section-title">Анализируемый продукт</span><div class="block-p">{p2_analyzable_html}</div>'
                    lifecycle_box_html += (
                        f'<span class="block-section-title">Исходы к концу периода ({end_period_weeks_str})</span>'
                        + f'<p class="block-p">{p2_outcomes_html}</p>'
                    )
                    if p2_other_popular_html:
                        lifecycle_box_html += f'<span class="block-section-title">Среди прочих категорий</span><p class="block-p">{p2_other_popular_html}</p>'
                    lifecycle_box_html += (
                        f'<span class="block-section-title">Полураспад анализируемого продукта</span>'
                        + f'<div class="block-p">{p3_html}</div>'
                    )
                    if p4_html:
                        lifecycle_box_html += f'<span class="block-section-title">Устойчивый перерыв и уход из анализируемого продукта</span><div class="block-p">{p4_html}</div>'
                    lifecycle_box_html += "</div></div>"

                    st.markdown(lifecycle_box_html, unsafe_allow_html=True)

                    # Таблица цикла жизни для Excel (из тех же данных, что и HTML-таблица)
                    lifecycle_table_rows = []
                    for t in range(k_int_lc):
                        period_label = f"{period_unit_single_lc} {t + 1}"
                        row = [period_label]
                        for b in BUCKET_ORDER:
                            cnt = int(bucket_counts_reindexed.loc[t, b]) if t in bucket_counts_reindexed.index else 0
                            pct = 100 * cnt / N_lc if N_lc else 0
                            row.append(f"{cnt} ({pct:.1f}%)")
                        lifecycle_table_rows.append(row)
                    lifecycle_table_columns = ["Период от когорты"] + [BUCKET_LABELS[b] for b in BUCKET_ORDER]
                    lifecycle_table_df = pd.DataFrame(lifecycle_table_rows, columns=lifecycle_table_columns)

                    # Структурированный вывод для Excel (заголовки, строки, отступы)
                    lifecycle_excel_rows = []
                    lifecycle_excel_rows.append(("heading", "Продажи анализируемого товара на объём якорного"))
                    if not cohort_clients_filtered:
                        lifecycle_excel_rows.append(("line", "В выбранных кластерах нет клиентов — коэффициент не рассчитан."))
                    elif q_anchor_lc and q_anchor_lc > 0:
                        lifecycle_excel_rows.append(("line", f"Объём анализируемого товара на единицу якорного товара: {r_ratio_lc:.2f}."))
                        analyzable_names_plain = ", ".join(selected_categories_lifecycle) if len(selected_categories_lifecycle) > 1 else (selected_categories_lifecycle[0] if selected_categories_lifecycle else "")
                        lifecycle_excel_rows.append((
                            "line",
                            f"При продаже {n_anchor_lc} ед. {category_label} в течении {k_periods_lifecycle} {period_word} будет продано {expected_int_lc} ед. {analyzable_names_plain}."
                        ))
                    else:
                        lifecycle_excel_rows.append(("line", "В выбранных кластерах и периоде нет покупок якорного товара — коэффициент не рассчитан."))

                    lifecycle_excel_rows.append(("heading", "Якорный продукт"))
                    lifecycle_excel_rows.append(("line", f"{period_unit_single} {week_1} — {pct_anchor_first:.1f}%"))
                    lifecycle_excel_rows.append(("line", f"Середина ({period_unit_single} {week_mid}) — {pct_anchor_mid:.1f}%"))
                    lifecycle_excel_rows.append(("line", f"К концу ({period_unit_single} {week_end}) — {pct_anchor_last:.1f}%"))
                    if p2_analyzable_lines:
                        lifecycle_excel_rows.append(("heading", "Анализируемый продукт"))
                        for frag in p2_analyzable_lines:
                            if "block-spacer" in frag:
                                lifecycle_excel_rows.append(("spacer", ""))
                            else:
                                lifecycle_excel_rows.append(("line", _html_to_plain_fragment(frag)))
                    lifecycle_excel_rows.append(("heading", f"Исходы к концу периода ({end_period_weeks_str})"))
                    lifecycle_excel_rows.append(("line", _html_to_plain_fragment(p2_outcomes_html)))
                    if p2_other_popular_html:
                        lifecycle_excel_rows.append(("heading", "Среди прочих категорий"))
                        lifecycle_excel_rows.append(("line", _html_to_plain_fragment(p2_other_popular_html)))
                    lifecycle_excel_rows.append(("heading", "Полураспад анализируемого продукта"))
                    for div in half_life_divs:
                        if "block-spacer" in div:
                            lifecycle_excel_rows.append(("spacer", ""))
                        else:
                            lifecycle_excel_rows.append(("line", _html_to_plain_fragment(div)))
                    if p4_html:
                        lifecycle_excel_rows.append(("heading", "Устойчивый перерыв и уход из анализируемого продукта"))
                        for line in p4_lines:
                            lifecycle_excel_rows.append(("line", _html_to_plain_fragment(line)))

                    # Формируем полный отчёт в Excel для кнопки скачивания (доступен после первого прохода по блокам)
                    cluster_summary_for_excel = st.session_state.get("report_cluster_summary")
                    cluster_comments_for_excel = st.session_state.get("report_cluster_comments", {})
                    html_without_css = _strip_css_from_html(lifecycle_box_html)
                    lifecycle_text = _html_to_plain_text(html_without_css)
                    safe_filename = "CLF " + re.sub(r'[*\\/:?"<>|]', "_", category_label) + ".xlsx"
                    st.session_state["excel_report_filename"] = safe_filename
                    try:
                        excel_bytes = build_excel_report(
                            cohort_start=cohort_start_global,
                            cohort_end=cohort_end_global,
                            anchor_product=category_label,
                            categories=selected_categories_global,
                            k_periods=int(k_periods_global),
                            is_months=is_months,
                            cluster_summary=cluster_summary_for_excel,
                            cluster_comments=cluster_comments_for_excel,
                            lifecycle_clusters=selected_clusters_lifecycle,
                            lifecycle_table=lifecycle_table_df,
                            lifecycle_output_text=lifecycle_text,
                            lifecycle_output_rows=lifecycle_excel_rows,
                        )
                        st.session_state["excel_report_bytes"] = excel_bytes
                        if not had_excel_bytes:
                            st.rerun()
                    except Exception:
                        st.session_state["excel_report_bytes"] = None

    else:
        st.warning("Загрузите оба документа в формате по шаблону (5 столбцов: категория, период, период, количество, код клиента).")
